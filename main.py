"""
AI Memory Gateway — 带记忆系统的 LLM 转发网关
=============================================
让你的 AI 拥有长期记忆。

工作原理：
1. 接收客户端（Kelivo / ChatBox / 任何 OpenAI 兼容客户端）的消息
2. 自动搜索数据库中的相关记忆，注入 system prompt
3. 转发给 LLM API（支持 OpenRouter / OpenAI / 任何兼容接口）
4. 后台自动存储对话 + 用 AI 提取新记忆

环境变量 MEMORY_ENABLED=false 时退化为纯转发网关（第一阶段）。
"""

import os
import json
import uuid
import asyncio
import httpx
from contextlib import asynccontextmanager
from fastapi import FastAPI, Request, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware

from database import (
    init_tables, close_pool, get_pool, save_message, search_memories, save_memory,
    get_all_memories_count, get_recent_memories, get_recent_conversation, delete_memory,
    clear_all_memories, update_memory, check_memory_duplicate,
    migrate_embeddings, get_embedding_stats,
    # v5.3 时间有效期 + 矛盾检测
    invalidate_memory, create_memory_edge, detect_contradictions,
    get_all_providers, get_provider, create_provider, update_provider, delete_provider,
    get_provider_models, get_all_saved_models, add_provider_model, update_provider_model, delete_provider_model,
    resolve_provider_for_model,
    get_all_categories, create_category, update_category, delete_category, match_category_by_name,
    get_system_prompt_from_db, set_system_prompt_in_db,
    # v4.1 云端同步
    sync_get_conversations, sync_get_conversation, sync_upsert_conversation, sync_delete_conversation,
    sync_upsert_messages, sync_get_projects, sync_upsert_project, sync_delete_project, sync_import_all,
    # v4.2 提醒系统
    create_reminder, get_reminders, update_reminder, delete_reminder, get_due_reminders, fire_reminder,
)
from config import (
    get_all_config, set_config, get_config, get_config_int, get_config_bool,
)
from memory_extractor import extract_memories
from mcp_server import get_mcp_app, get_calendar_mcp_app, mcp_memory, mcp_calendar
from web_search import web_search, format_results_for_prompt, get_engine_list
from mcp_client import get_tools_for_servers, run_tool_call_loop, call_tool, call_tools_batch, clear_tool_cache

# ============================================================
# 配置项 —— 全部从环境变量读取，部署时在云平台面板里设置
# ============================================================

# 你的 API Key（OpenRouter / OpenAI / 其他兼容服务）
API_KEY = os.getenv("API_KEY", "")

# API 地址（改这个就能切换不同的 LLM 服务商）
API_BASE_URL = os.getenv("API_BASE_URL", "https://openrouter.ai/api/v1/chat/completions")

# 默认模型（如果客户端没指定就用这个）
DEFAULT_MODEL = os.getenv("DEFAULT_MODEL", "anthropic/claude-sonnet-4")

# 网关端口
PORT = int(os.getenv("PORT", "8080"))

# 记忆系统开关（数据库出问题时可以临时关掉）
MEMORY_ENABLED = os.getenv("MEMORY_ENABLED", "false").lower() == "true"

# 每次注入的最大记忆条数
MAX_MEMORIES_INJECT = int(os.getenv("MAX_MEMORIES_INJECT", "15"))

# 记忆提取间隔：每隔几轮对话提取一次记忆（默认3轮）
MEMORY_EXTRACT_INTERVAL = int(os.getenv("MEMORY_EXTRACT_INTERVAL", "3"))

# 前端访问密码（不设就不需要密码）
ACCESS_TOKEN = os.getenv("ACCESS_TOKEN", "")


# ============================================================
# 动态配置读取（v3.1）
# ============================================================
# 配置优先级：数据库 > 环境变量 > 默认值
# 以上三个变量保留作为启动时/数据库不可用时的降级值
# 运行时通过以下函数读取最新配置

async def get_memory_enabled() -> bool:
    """读取记忆开关（动态）"""
    try:
        return await get_config_bool("memory_enabled", fallback=MEMORY_ENABLED)
    except Exception:
        return MEMORY_ENABLED

async def get_max_inject() -> int:
    """读取注入条数（动态）"""
    try:
        return await get_config_int("max_inject", fallback=MAX_MEMORIES_INJECT)
    except Exception:
        return MAX_MEMORIES_INJECT

async def get_extract_interval() -> int:
    """读取提取间隔（动态）"""
    try:
        return await get_config_int("extract_interval", fallback=MEMORY_EXTRACT_INTERVAL)
    except Exception:
        return MEMORY_EXTRACT_INTERVAL

# 额外的请求头（有些 API 需要，比如 OpenRouter 需要 Referer）
EXTRA_REFERER = os.getenv("EXTRA_REFERER", "https://ai-memory-gateway.local")
EXTRA_TITLE = os.getenv("EXTRA_TITLE", "AI Memory Gateway")






# ============================================================
# 对话计数器（控制记忆提取频率）
# ============================================================

_conversation_counter = 0
_counter_lock = asyncio.Lock()


# ============================================================
# 后台任务引用（防止 GC 回收）
# ============================================================

_background_tasks: set = set()


def _spawn_background_task(coro):
    """启动后台任务并保留引用，避免在执行中被 GC 回收。"""
    task = asyncio.create_task(coro)
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)
    return task


# ============================================================
# 人设加载
# ============================================================

def load_system_prompt():
    """从 system_prompt.txt 文件读取人设内容（降级方案）"""
    prompt_path = os.path.join(os.path.dirname(__file__), "system_prompt.txt")
    try:
        with open(prompt_path, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if content:
                return content
    except FileNotFoundError:
        pass
    print("ℹ️  未找到 system_prompt.txt 或文件为空，将不注入 system prompt")
    return ""


# 文件版作为启动降级值
_FILE_SYSTEM_PROMPT = load_system_prompt()
# 运行时变量（可被数据库覆盖）
SYSTEM_PROMPT = _FILE_SYSTEM_PROMPT

if SYSTEM_PROMPT:
    print(f"✅ 人设已加载（文件），长度：{len(SYSTEM_PROMPT)} 字符")
else:
    print("ℹ️  无人设，纯转发模式")


async def get_active_system_prompt() -> str:
    """获取当前生效的 system prompt（数据库优先，文件降级）"""
    try:
        db_prompt = await get_system_prompt_from_db()
        if db_prompt is not None:
            return db_prompt
    except Exception:
        pass
    return _FILE_SYSTEM_PROMPT


# ============================================================
# 应用生命周期管理
# ============================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用启动时初始化数据库和MCP，关闭时断开连接"""
    digest_task = None
    dream_check_task = None
    
    if MEMORY_ENABLED:
        try:
            await init_tables()
            count = await get_all_memories_count()
            print(f"✅ 记忆系统已启动，当前记忆数量：{count}")
            print(f"📊 记忆提取间隔：每 {MEMORY_EXTRACT_INTERVAL} 轮对话提取一次")
            
            # v5.6：首次启动时将出厂默认 prompt 写入 config 表（空值才写入）
            try:
                factory = _get_factory_prompts()
                seeded = 0
                for key, default_text in factory.items():
                    existing = await get_config(key)
                    if not existing:
                        await set_config(key, default_text)
                        seeded += 1
                if seeded > 0:
                    print(f"📝 首次启动：写入了 {seeded} 个默认 prompt 到配置表")
            except Exception as e:
                print(f"⚠️  默认 prompt 初始化失败: {e}")
            
            # 启动每日记忆整理调度器
            from daily_digest import daily_digest_scheduler
            digest_task = asyncio.create_task(daily_digest_scheduler())
            
            # 启动自动 Dream 检查器（每小时检查24h无活动）
            from dream import auto_dream_scheduler
            dream_check_task = asyncio.create_task(auto_dream_scheduler())
            
        except Exception as e:
            print(f"⚠️  数据库初始化失败: {e}")
            print("⚠️  记忆系统将不可用，但网关仍可正常转发")
    else:
        print("ℹ️  记忆系统已关闭（设置 MEMORY_ENABLED=true 开启）")
    
    # 启动 MCP session managers（v5.4：两个模块）
    async with mcp_memory.session_manager.run():
        async with mcp_calendar.session_manager.run():
            print("✅ MCP server 已启动（/memory/mcp + /calendar/mcp）")
            yield
    
    if digest_task:
        digest_task.cancel()
    if dream_check_task:
        dream_check_task.cancel()
    if MEMORY_ENABLED:
        await close_pool()


app = FastAPI(title="AI Memory Gateway", version="3.1.0", lifespan=lifespan)


# ============================================================
# Admin 认证中间件 — 保护 /admin/* 和 /debug/* 端点
# ============================================================

class AdminAuthMiddleware(BaseHTTPMiddleware):
    """当设置了 ACCESS_TOKEN 时，需要认证的路径前缀。

    - /admin/、/debug/、/sync/：管理面板 / 调试 / 云同步导出导入
    - /projects/：v5.8 项目文件上传与 chunk 管理（POST /projects/{id}/files/{fid}/process
      与 DELETE /projects/{id}/files/{fid}/chunks），写操作必须保护
    - /import/：导入种子记忆（如 /import/seed-memories），会写入数据库
    """

    PROTECTED_PREFIXES = ("/admin/", "/debug/", "/sync/", "/projects/", "/import/")
    
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        
        # OPTIONS 预检请求直接放行
        if request.method == "OPTIONS":
            return await call_next(request)
        
        # /admin 面板页面本身不拦截（不带斜杠的精确匹配）
        if path == "/admin":
            return await call_next(request)
        
        # 只拦截受保护路径
        if ACCESS_TOKEN and any(path.startswith(p) for p in self.PROTECTED_PREFIXES):
            # 从 Authorization header 或 query param 读 token
            auth = request.headers.get("Authorization", "")
            token = auth.replace("Bearer ", "") if auth.startswith("Bearer ") else ""
            if not token:
                token = request.query_params.get("token", "")
            
            if token != ACCESS_TOKEN:
                return JSONResponse(
                    status_code=401,
                    content={"error": "未授权访问，请提供有效的 ACCESS_TOKEN"}
                )
        
        return await call_next(request)


app.add_middleware(AdminAuthMiddleware)

# ============================================================
# CORS 配置 — 从环境变量读取允许的域名
# ============================================================

# CORS 白名单：通过环境变量配置，逗号分隔
# 示例：CORS_ORIGINS=https://your-frontend.example.com,http://localhost:5173
_cors_env = os.getenv("CORS_ORIGINS", "http://localhost:5173,http://localhost:5174")
_cors_origins = [o.strip() for o in _cors_env.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# 模板变量替换
# ============================================================

from datetime import datetime, timezone, timedelta

TZ_CST = timezone(timedelta(hours=8))  # 东八区

def replace_template_variables(text: str, context: dict = None) -> str:
    """
    替换 system prompt / skill prompt 中的模板变量。
    支持的变量：
      {cur_datetime}    → 2026-03-24 14:30:00
      {cur_date}        → 2026-03-24
      {cur_time}        → 14:30:00
      {cur_weekday}     → 星期一
      {model_name}      → deepseek/deepseek-chat-v3-0324
      {user_name}       → 用户昵称
      {assistant_name}  → AI名字
    """
    if not text or '{' not in text:
        return text

    now = datetime.now(TZ_CST)
    weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
    ctx = context or {}

    replacements = {
        '{cur_datetime}':   now.strftime('%Y-%m-%d %H:%M:%S'),
        '{cur_date}':       now.strftime('%Y-%m-%d'),
        '{cur_time}':       now.strftime('%H:%M:%S'),
        '{cur_weekday}':    weekdays[now.weekday()],
        '{model_name}':     ctx.get('model_name', ''),
        '{user_name}':      ctx.get('user_name', ''),
        '{assistant_name}': ctx.get('assistant_name', ''),
    }

    for key, val in replacements.items():
        # 即便 val 是空字符串也要替换，否则 {user_name} 等占位符会原样残留进 prompt
        if key in text and isinstance(val, str):
            text = text.replace(key, val)

    return text


# ============================================================
# 记忆注入
# ============================================================

async def build_system_prompt_with_memories(user_message: str, user_msg_count: int = 1, project_id: str = None) -> tuple:
    """
    构建带记忆的 system prompt（v5.5 日历层级注入 + v5.8 项目注入 + 缓存优化）
    
    返回 (prompt_str, metadata_dict)，metadata 可能包含 handoff 信息。
    
    注入顺序（为 Prompt Caching 优化——不变的在前，变化的在后）：
    ── 静态区（同一天内不变，命中缓存 1/10 价格）──
    1. 人设
    2. 用户画像
    3. 锁定记忆（很少变）
    4. 日历层级注入（一天内不变）
    5. 项目指令（静态，整个项目内不变）
    ── 动态区（每轮变化，不缓存）──
    6. 语义搜索碎片（每轮根据用户消息重新搜索，含项目记忆）
    7. 项目文件相关片段（语义搜索）
    8. Dream 犯困提示
    """
    active_prompt = await get_active_system_prompt()
    prompt_meta = {}
    
    # ---- ① 用户画像（静态，一天不变）----
    try:
        user_profile = await get_config("user_profile")
        if user_profile:
            active_prompt += f"\n\n【用户画像】\n{user_profile}"
    except Exception as e:
        print(f"⚠️  用户画像读取失败: {e}")
    
    if not await get_memory_enabled():
        return active_prompt, prompt_meta
    
    try:
        # ---- ② 锁定记忆：全量注入（静态，很少变）----
        from database import get_permanent_memories
        permanent = await get_permanent_memories()
        if permanent:
            perm_lines = []
            for mem in permanent:
                title = mem.get("title", "")
                content = mem.get("content", "")
                if title:
                    perm_lines.append(f"- 【{title}】{content}")
                else:
                    perm_lines.append(f"- {content}")
            perm_text = "\n".join(perm_lines)
            active_prompt += f"\n\n【长期记忆（用户标记为重要）】\n{perm_text}"
            print(f"📌 注入了 {len(permanent)} 条锁定记忆")
        
        # ---- ③ 日历层级注入（静态，一天内不变）----
        try:
            calendar_enabled = await get_config("calendar_inject_enabled")
            if calendar_enabled is None or str(calendar_enabled).lower() != 'false':
                from database import get_calendar_for_injection
                cal_entries = await get_calendar_for_injection(lookback_days=365)
                if cal_entries:
                    cal_lines = []
                    for entry in cal_entries:
                        label = entry.get("label", "")
                        # 优先用 digest（模型注入版），没有就用 summary（兜底），都没有就跳过
                        text = entry.get("digest") or entry.get("summary") or ""
                        if text:
                            cal_lines.append(f"📅 {label}：{text}")
                    if cal_lines:
                        cal_text = "\n".join(cal_lines)
                        active_prompt += f"\n\n【近期日历（从大到小的层级记忆，越远越概括）】\n{cal_text}"
                        print(f"📅 日历注入了 {len(cal_lines)} 条层级记忆")
        except Exception as e:
            print(f"⚠️  日历注入失败: {e}")
        
        # ---- ④ 项目指令注入（静态，整个项目内不变）----
        if project_id:
            try:
                from database import get_project_by_id
                proj = await get_project_by_id(project_id)
                if proj and proj.get("instructions"):
                    active_prompt += f"\n\n【项目指令】\n{proj['instructions']}"
                    print(f"📂 注入了项目指令（项目: {proj.get('name', project_id)}）")
            except Exception as e:
                print(f"⚠️  项目指令注入失败: {e}")
        
        # ---- 静态/动态分隔标记（用于 Prompt Caching）----
        # 上面的人设+画像+锁定记忆+日历是静态的（一天内不变），下面的搜索碎片/犯困/切窗是动态的
        active_prompt += "\n\n<!-- CACHE_BOUNDARY -->"
        
        # ---- ⑤ 语义搜索碎片（动态，每轮变化）----
        inject_limit = await get_max_inject()
        memories = await search_memories(user_message, limit=inject_limit, project_id=project_id)
        
        # 加载热度参数（v5.4：可配置阈值）
        from database import get_heat_params
        heat_params = await get_heat_params()
        th_high = heat_params["threshold_high"]
        th_medium = heat_params["threshold_medium"]
        
        # v5.6：中热度摘要截断字数（可配置）
        truncate_len = await get_config_int("heat_medium_truncate", fallback=60)
        
        # 过滤掉已经在永久记忆里注入过的
        perm_ids = {m["id"] for m in permanent} if permanent else set()
        memories = [m for m in memories if m.get("id") not in perm_ids]
        
        if memories:
            memory_lines = []
            for mem in memories:
                title = mem.get("title", "")
                heat = mem.get("heat", 1.0)
                date_tag = ""
                if mem.get("created_at"):
                    try:
                        from datetime import datetime
                        dt = mem["created_at"]
                        if hasattr(dt, "strftime"):
                            date_tag = f"[{dt.strftime('%Y-%m-%d')}]"
                        else:
                            date_tag = f"[{str(dt)[:10]}]"
                    except Exception:
                        pass
                
                cat_name = mem.get("category_name", "")
                cat_tag = f"({cat_name})" if cat_name else ""
                
                # v5.4 热度分档注入（阈值可配置）
                if heat > th_high:
                    if title:
                        memory_lines.append(f"- {date_tag}{cat_tag}【{title}】{mem['content']}")
                    else:
                        memory_lines.append(f"- {date_tag}{cat_tag} {mem['content']}")
                elif heat > th_medium:
                    if title:
                        brief = mem['content'][:truncate_len] + "…" if len(mem['content']) > truncate_len else mem['content']
                        memory_lines.append(f"- {date_tag}{cat_tag}【{title}】{brief}（印象模糊）")
                    else:
                        brief = mem['content'][:truncate_len] + "…" if len(mem['content']) > truncate_len else mem['content']
                        memory_lines.append(f"- {date_tag}{cat_tag} {brief}（印象模糊）")
                
            memory_text = "\n".join(memory_lines)
            
            if memory_lines:
                active_prompt += f"\n\n【从过往对话中检索到的相关记忆】\n以下是与当前话题可能相关的历史信息，自然地融入对话中，不要刻意提起'我记得'：\n{memory_text}"
                skipped = len(memories) - len(memory_lines)
                skip_msg = f"（跳过 {skipped} 条低热度）" if skipped > 0 else ""
                print(f"📚 注入了 {len(memory_lines)} 条相关记忆{skip_msg}（热度分档注入）")
            else:
                print(f"📚 搜到 {len(memories)} 条记忆但全部热度过低，跳过注入")
        
        # ---- ⑥ 项目文件相关片段（动态，每轮根据用户消息搜索）----
        if project_id:
            try:
                from database import search_file_chunks
                file_chunks = await search_file_chunks(project_id, user_message, limit=6)
                if file_chunks:
                    chunk_lines = []
                    for chunk in file_chunks:
                        chunk_lines.append(f"📎 [{chunk['file_name']}] {chunk['content']}")
                    chunk_text = "\n".join(chunk_lines)
                    active_prompt += f"\n\n【项目文件中的相关内容】\n{chunk_text}"
                    print(f"📂 注入了 {len(file_chunks)} 条文件片段")
            except Exception as e:
                print(f"⚠️  文件搜索失败: {e}")
        
        # ---- ⑦ Dream 犯困提示（动态）----
        try:
            from dream import get_drowsy_prompt
            drowsy = await get_drowsy_prompt()
            if drowsy:
                active_prompt += f"\n{drowsy}"
                print(f"😴 注入了犯困提示")
        except Exception:
            pass
        
        # ---- ⑥ 无缝切窗（动态，仅前几轮）----
        try:
            handoff_on = await get_config_bool("handoff_enabled", fallback=True)
            handoff_stop = await get_config_int("handoff_stop_rounds", fallback=3)
            if handoff_on and user_msg_count <= handoff_stop:
                from database import get_handoff_messages
                handoff_count = await get_config_int("handoff_msg_count", fallback=6)
                handoff_msgs, prev_title = await get_handoff_messages(limit=handoff_count)
                if handoff_msgs:
                    title_hint = f"（上一个对话：{prev_title}）" if prev_title else ""
                    prompt_meta["handoff"] = {"title": prev_title or "", "count": len(handoff_msgs)}
                    
                    if user_msg_count == 1:
                        # 第 1 轮：注入原文消息（完整上下文做自然衔接）
                        handoff_lines = []
                        for m in handoff_msgs:
                            role_label = "用户" if m["role"] == "user" else "助手"
                            content = m.get("content", "")
                            if len(content) > 500:
                                content = content[:500] + "…（截断）"
                            handoff_lines.append(f"{role_label}: {content}")
                        handoff_text = "\n".join(handoff_lines)
                        active_prompt += f"\n\n【上一个对话的最近内容{title_hint}】\n以下是用户在上一个对话窗口最后聊的内容，自然衔接即可，不要说'我看到你上次聊了'：\n{handoff_text}"
                        print(f"🔗 无缝切窗：注入了 {len(handoff_msgs)} 条原文消息（第 1/{handoff_stop} 轮）")
                        
                        # 后台异步生成摘要，供第 2 轮起使用
                        # 用 prev_title 当 conv_id 的近似标识（避免额外查 conv_id）
                        _spawn_background_task(_generate_handoff_summary(
                            prev_title or "unknown", handoff_msgs, prev_title
                        ))
                    else:
                        # 第 2+ 轮：优先使用缓存的摘要
                        cached = _handoff_summary_cache.get("summary")
                        if cached:
                            active_prompt += f"\n\n【上一个对话摘要{title_hint}】\n{cached}"
                            print(f"🔗 无缝切窗：注入摘要（{len(cached)}字，第 {user_msg_count}/{handoff_stop} 轮）")
                        else:
                            # 摘要还没生成好，降级用原文
                            handoff_lines = []
                            for m in handoff_msgs:
                                role_label = "用户" if m["role"] == "user" else "助手"
                                content = m.get("content", "")
                                if len(content) > 500:
                                    content = content[:500] + "…（截断）"
                                handoff_lines.append(f"{role_label}: {content}")
                            handoff_text = "\n".join(handoff_lines)
                            active_prompt += f"\n\n【上一个对话的最近内容{title_hint}】\n以下是用户在上一个对话窗口最后聊的内容，自然衔接即可，不要说'我看到你上次聊了'：\n{handoff_text}"
                            print(f"🔗 无缝切窗：摘要未就绪，降级注入原文（第 {user_msg_count}/{handoff_stop} 轮）")
        except Exception as e:
            print(f"⚠️  无缝切窗失败: {e}")
        
        return active_prompt, prompt_meta
        
    except Exception as e:
        print(f"⚠️  记忆检索失败: {e}，使用纯人设")
        return active_prompt, prompt_meta


# ============================================================
# 后台记忆处理
# ============================================================

# ============================================================
# 情绪检测（v5.2 热度系统）
# ============================================================

# 用户消息中的情绪关键词（规则引擎兜底）
EMOTION_HIGH_KEYWORDS = [
    "抱抱", "贴贴", "亲亲", "呜", "哭", "崩溃", "撑不住", "好难过",
    "好开心", "好幸福", "我爱你", "谢谢你", "你真好", "好想你",
    "对不起", "害怕", "不想活", "好累", "受不了", "心疼",
    "太好了", "我好高兴", "感动", "哭了",
]
EMOTION_MEDIUM_KEYWORDS = [
    "难过", "开心", "紧张", "焦虑", "生气", "委屈", "失落",
    "高兴", "感谢", "抱歉", "担心", "烦", "郁闷", "不舒服",
]


def detect_emotion_from_user_msg(text: str) -> str:
    """从用户消息检测情绪级别（规则引擎）"""
    if not text:
        return "normal"
    for kw in EMOTION_HIGH_KEYWORDS:
        if kw in text:
            return "high"
    for kw in EMOTION_MEDIUM_KEYWORDS:
        if kw in text:
            return "medium"
    return "normal"


def detect_emotion_from_response(text: str) -> str:
    """从模型回复中解析隐藏的情绪标记 <!--emotion:高-->"""
    if not text:
        return "normal"
    import re
    match = re.search(r'<!--\s*emotion\s*[:：]\s*(高|中|low|medium|high)\s*-->', text)
    if match:
        level = match.group(1)
        if level in ("高", "high"):
            return "high"
        elif level in ("中", "medium"):
            return "medium"
    return "normal"


def merge_emotion_levels(user_level: str, response_level: str) -> str:
    """取两个情绪级别的高值"""
    order = {"normal": 0, "medium": 1, "high": 2}
    return max([user_level, response_level], key=lambda x: order.get(x, 0))


def emotion_to_weight(level: str) -> int:
    """情绪级别转数字权重（0-10）"""
    return {"high": 8, "medium": 4, "normal": 0}.get(level, 0)


# 主动记忆触发词 —— 用户说了这些词就立刻提取，不等计数器
MEMORY_TRIGGER_WORDS = ["记住", "记下", "帮我记", "请记", "别忘了", "不要忘记", "你要记得", "记一下"]


async def process_memories_background(session_id: str, user_msg: str, assistant_msg: str, model: str, emotion_level: str = "normal", project_id: str = None):
    """
    后台异步：存储对话 + 按间隔提取记忆（不阻塞主流程）
    
    v2.4 改进：
    - 提取前的对比范围改为「搜索相关 + 最近记忆」组合，覆盖种子记忆
    - 存储前逐条做去重检测，防止冗余写入
    
    v2.5 改进：
    - 检测主动记忆触发词，命中时立即提取，不等计数器
    - 不重置计数器，不干扰正常的定时提取节奏
    
    v3.7 改进：
    - 提取时从数据库捞最近 N 轮完整对话（而不是只看最后一轮）
    - N = extract_interval，攒几轮就提取几轮
    
    v5.2 改进：
    - 接受 emotion_level 参数，传递给 save_memory 的 emotional_weight
    
    v5.8 改进：
    - 接受 project_id 参数，项目内对话提取的记忆自动打上 project_id 标签
    """
    global _conversation_counter

    try:
        # 对话始终保存
        await save_message(session_id, "user", user_msg, model)
        await save_message(session_id, "assistant", assistant_msg, model)

        # v5.4：检测用户是否让 AI 去睡觉（触发 Dream）
        _DREAM_TRIGGER_WORDS = ["去睡吧", "去睡觉", "睡一下", "去做梦", "去休息", "快去睡"]
        if any(kw in user_msg for kw in _DREAM_TRIGGER_WORDS):
            print(f"🌙 检测到睡觉指令，后台触发 Dream...")
            async def _silent_dream():
                try:
                    from dream import run_dream
                    async for event in run_dream(trigger_type="manual"):
                        if event["type"] == "error":
                            print(f"   🌙 Dream 出错: {event['data']}")
                        elif event["type"] == "complete":
                            print(f"   🌙 Dream 完成: {event['data']}")
                except Exception as e:
                    print(f"   🌙 Dream 异常: {e}")
            _spawn_background_task(_silent_dream())

        # 检测用户是否主动要求记忆
        force_extract = any(kw in user_msg for kw in MEMORY_TRIGGER_WORDS)

        # 使用锁保护计数器，防止并发请求导致重复提取或跳过
        should_extract = False
        async with _counter_lock:
            _conversation_counter += 1
            extract_interval = await get_extract_interval()
            if _conversation_counter < extract_interval and not force_extract:
                print(f"💬 对话已保存（{_conversation_counter}/{extract_interval}轮后提取记忆）")
                return {"action": "skip", "counter": _conversation_counter, "interval": extract_interval}

            if force_extract:
                print(f"🎯 检测到主动记忆请求，立即提取（计数器保持 {_conversation_counter}/{extract_interval}）")
            else:
                _conversation_counter = 0
                print(f"🧠 达到提取间隔（{extract_interval}轮），开始提取记忆...")
            should_extract = True
        
        # ===== v2.4 改进：组合式获取已有记忆 =====
        # 用当前对话内容搜索相关记忆（能覆盖到种子记忆）
        # track_recall=False: 这里是去重对比，不是用户聊天，不应该增加召回计数
        related = await search_memories(user_msg, limit=50, track_recall=False, project_id=project_id)
        related_contents = [r["content"] for r in related]
        
        # 再补充最近的记忆（防止遗漏新存的）
        recent = await get_recent_memories(limit=30, project_id=project_id)
        recent_contents = [r["content"] for r in recent]
        
        # 合并去重
        seen = set()
        existing_contents = []
        for content in related_contents + recent_contents:
            if content not in seen:
                seen.add(content)
                existing_contents.append(content)
        
        print(f"📋 对比范围：{len(existing_contents)} 条已有记忆（搜索相关 {len(related_contents)} + 最近 {len(recent_contents)}，去重后 {len(existing_contents)}）")
        
        # ===== v3.7 改进：攒 N 轮完整对话一起提取 =====
        # 从数据库捞最近 N*2 条消息（N轮 = N条user + N条assistant）
        recent_msgs = await get_recent_conversation(limit=extract_interval * 2)
        
        if recent_msgs:
            messages_for_extraction = [
                {"role": row["role"], "content": row["content"]}
                for row in recent_msgs
            ]
            print(f"📨 提取范围：最近 {len(messages_for_extraction)} 条消息（约 {len(messages_for_extraction)//2} 轮对话）")
        else:
            # 降级：如果数据库查不到，至少用当前这一轮
            messages_for_extraction = [
                {"role": "user", "content": user_msg},
                {"role": "assistant", "content": assistant_msg},
            ]
            print(f"📨 提取范围：当前 1 轮对话（降级）")
        
        # 获取可用分类名（用于自动归类）
        try:
            all_cats = await get_all_categories()
            cat_names = [c["name"] for c in all_cats]
        except Exception:
            cat_names = []
        
        # 读取数据库中的模型和提示词配置
        from config import get_config
        db_memory_model = await get_config("default_memory_model")
        db_memory_prompt = await get_config("prompt_memory_extract")
        
        new_memories = await extract_memories(
            messages_for_extraction,
            existing_memories=existing_contents,
            categories=cat_names,
            model_override=db_memory_model if db_memory_model else None,
            prompt_override=db_memory_prompt if db_memory_prompt else None,
            emotion_level=emotion_level,
        )
        
        # 过滤垃圾记忆（不靠模型自觉，硬过滤）
        META_BLACKLIST = [
            "记忆库", "记忆系统", "检索", "没有被记录", "没有被提取",
            "记忆遗漏", "尚未被记录", "写入不完整", "检索功能",
            "系统没有返回", "关键词匹配", "语义匹配", "语义检索",
            "阈值", "数据库", "seed", "导入", "部署",
            "bug", "debug", "端口", "网关",
        ]
        
        filtered_memories = []
        for mem in new_memories:
            content = mem["content"]
            if any(kw in content for kw in META_BLACKLIST):
                print(f"🚫 过滤掉meta记忆: {content[:60]}...")
                continue
            filtered_memories.append(mem)
        
        # ===== v5.3 改进：去重 + 矛盾检测（共用一次搜索）=====
        saved_count = 0
        skipped_count = 0
        contradiction_count = 0
        
        for mem in filtered_memories:
            # 去重检测（v5.4：传标题，标题不同时不误杀）
            is_dup, similar_results = await check_memory_duplicate(mem["content"], new_title=mem.get("title", ""))
            
            if is_dup:
                skipped_count += 1
                continue
            
            # 矛盾检测（v5.3：复用去重搜索结果，不额外调 embedding API）
            contradicted_ids = detect_contradictions(
                mem.get("title", ""), mem["content"], similar_results
            )
            
            # 自动匹配分类
            cat_id = None
            cat_hint = mem.get("category", "")
            if cat_hint:
                cat_id = await match_category_by_name(cat_hint)
            
            # 保存新记忆（v5.3：返回 ID，用于创建 supersedes edge）
            new_id = await save_memory(
                content=mem["content"],
                importance=mem["importance"],
                source_session=session_id,
                title=mem.get("title", ""),
                category_id=cat_id,
                source="ai_extracted",
                emotional_weight=mem.get("emotional_weight", 0) or emotion_to_weight(emotion_level),
                project_id=project_id,
            )
            saved_count += 1
            
            # 处理矛盾：标旧记忆失效 + 创建 supersedes edge
            if contradicted_ids and new_id:
                for old_id in contradicted_ids:
                    await invalidate_memory(old_id, reason=f"被新记忆 #{new_id} 替代")
                    await create_memory_edge(
                        new_id, "memory", old_id, "memory", "supersedes",
                        reason="提取时自动检测到信息更新", created_by="extractor"
                    )
                    contradiction_count += 1
        
        if saved_count > 0 or skipped_count > 0:
            total = await get_all_memories_count()
            contra_msg = f"，{contradiction_count} 条旧记忆被替代" if contradiction_count > 0 else ""
            print(f"💾 提取结果：{saved_count} 条新记忆已保存，{skipped_count} 条重复已跳过{contra_msg}，总计 {total} 条")
            return {"action": "extract", "saved": saved_count, "skipped": skipped_count, "contradictions": contradiction_count, "total": total}
        else:
            print(f"💭 本轮对话未产生新记忆")
            return {"action": "extract", "saved": 0, "skipped": 0, "contradictions": 0, "total": await get_all_memories_count()}
            
    except Exception as e:
        print(f"⚠️  后台记忆处理失败: {e}")
        return {"action": "error", "error": str(e)}


# ============================================================
# API 接口
# ============================================================

@app.get("/")
async def root_status():
    """根路由 — 返回系统状态 JSON（admin面板依赖此接口获取统计数据）"""
    memory_count = 0
    mem_enabled = await get_memory_enabled()
    if mem_enabled:
        try:
            memory_count = await get_all_memories_count()
        except Exception:
            pass
    return {
        "status": "running",
        "gateway": "AI Memory Gateway v3.1 (动态配置)",
        "version": "AI Memory Gateway v3.1",
        "memory_enabled": mem_enabled,
        "memory_count": memory_count,
        # 前端 admin-panel 读 status.memories, 加别名避免显示 '-'
        "memories": memory_count,
        "max_inject": await get_max_inject(),
        "default_model": DEFAULT_MODEL,
        "extract_interval": await get_extract_interval(),
    }


@app.get("/favicon.ico")
async def favicon():
    """返回空favicon防止404"""
    return Response(status_code=204)


@app.post("/auth/verify")
async def auth_verify(request: Request):
    """验证前端访问密码"""
    if not ACCESS_TOKEN:
        return {"status": "ok", "message": "无需密码"}
    try:
        data = await request.json()
        token = data.get("token", "")
        if token == ACCESS_TOKEN:
            return {"status": "ok"}
        return JSONResponse(status_code=401, content={"error": "密码错误"})
    except Exception:
        return JSONResponse(status_code=401, content={"error": "密码错误"})


@app.get("/api/status")
async def api_status():
    """系统状态（JSON）— 兼容旧接口，重定向到根路由"""
    return await root_status()


@app.get("/admin")
async def admin_panel():
    """返回管理面板 HTML（admin-panel/index.html）。

    v1.1 重构时这里被改成了返回 JSON, 直接访问 /admin 看到 {"status": "running"}
    而不是面板页面。改回 FileResponse, 文件不存在时降级返回 JSON。
    """
    from fastapi.responses import FileResponse
    panel_path = os.path.join(os.path.dirname(__file__), "admin-panel", "index.html")
    if os.path.exists(panel_path):
        return FileResponse(panel_path, media_type="text/html; charset=utf-8")
    return {"status": "running", "service": "kiwi-mem", "warning": "admin-panel/index.html not found"}


@app.get("/v1/models")
async def list_models():
    """从 OpenRouter 拉取完整模型列表"""
    try:
        # 从 API_BASE_URL 提取基础地址（去掉 /chat/completions 部分）
        base = API_BASE_URL.split("/chat/completions")[0].rstrip("/")
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                f"{base}/models",
                headers={"Authorization": f"Bearer {API_KEY}"},
            )
            if resp.status_code == 200:
                return resp.json()
    except Exception as e:
        print(f"⚠️ 拉取模型列表失败: {e}")
    # 失败时返回默认模型兜底
    return {
        "object": "list",
        "data": [
            {
                "id": DEFAULT_MODEL,
                "object": "model",
                "created": 1700000000,
                "owned_by": "ai-memory-gateway",
            }
        ],
    }


# ============================================================
# 文件内容提取
# ============================================================

@app.post("/v1/files/extract")
async def extract_file_content(file: UploadFile = File(...)):
    """提取上传文件的文本内容（PDF/DOCX/XLSX/ZIP等）"""
    import io
    import tempfile
    
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    content_bytes = await file.read()
    
    try:
        extracted = ""
        file_type = ext
        
        # PDF
        if ext == "pdf":
            try:
                from pypdf import PdfReader
                reader = PdfReader(io.BytesIO(content_bytes))
                pages = []
                for i, page in enumerate(reader.pages):
                    text = page.extract_text()
                    if text and text.strip():
                        pages.append(f"[第{i+1}页]\n{text.strip()}")
                extracted = "\n\n".join(pages) if pages else "(PDF 无法提取文字，可能是扫描件)"
            except Exception as e:
                extracted = f"(PDF 解析失败: {str(e)})"
        
        # Word DOCX
        elif ext == "docx":
            try:
                from docx import Document
                doc = Document(io.BytesIO(content_bytes))
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                extracted = "\n\n".join(paragraphs) if paragraphs else "(DOCX 无内容)"
            except Exception as e:
                extracted = f"(DOCX 解析失败: {str(e)})"
        
        # Excel XLSX
        elif ext in ("xlsx", "xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                        if row_str.strip():
                            rows.append(row_str)
                    if rows:
                        sheets.append(f"[工作表: {sheet_name}]\n" + "\n".join(rows[:500]))  # 限制行数
                extracted = "\n\n".join(sheets) if sheets else "(XLSX 无内容)"
                wb.close()
            except Exception as e:
                extracted = f"(XLSX 解析失败: {str(e)})"
        
        # ZIP — 列出文件列表 + 提取文本文件内容
        elif ext == "zip":
            import zipfile
            try:
                zf = zipfile.ZipFile(io.BytesIO(content_bytes))
                file_list = zf.namelist()
                text_extensions = {'.txt', '.md', '.py', '.js', '.jsx', '.ts', '.tsx', '.json', '.csv', '.html', '.css', '.xml', '.yaml', '.yml', '.toml', '.sh', '.sql', '.java', '.c', '.cpp', '.go', '.rs', '.rb', '.log', '.ini', '.cfg', '.env'}
                
                parts = [f"压缩包共 {len(file_list)} 个文件：\n" + "\n".join(f"  {f}" for f in file_list[:100])]
                
                # 提取小的文本文件
                for name in file_list[:20]:
                    name_lower = name.lower()
                    if any(name_lower.endswith(e) for e in text_extensions):
                        info = zf.getinfo(name)
                        if info.file_size < 50000:  # 小于50KB
                            try:
                                text = zf.read(name).decode("utf-8", errors="ignore")
                                parts.append(f"\n[文件: {name}]\n{text}")
                            except Exception:
                                pass
                
                extracted = "\n".join(parts)
                zf.close()
            except Exception as e:
                extracted = f"(ZIP 解析失败: {str(e)})"
        
        # 其他文本类格式尝试直接读
        else:
            try:
                extracted = content_bytes.decode("utf-8", errors="ignore")
                if not extracted.strip() or '\x00' in extracted[:200]:
                    extracted = f"(二进制文件，无法提取文本内容)"
                    file_type = "binary"
            except Exception:
                extracted = f"(无法解析该文件格式)"
                file_type = "binary"
        
        # 截断过长的内容
        if len(extracted) > 100000:
            extracted = extracted[:100000] + f"\n\n...(内容过长，已截断至约10万字符)"
        
        return {
            "filename": filename,
            "type": file_type,
            "size": len(content_bytes),
            "text": extracted,
        }
    
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"error": f"文件处理失败: {str(e)}"}
        )


@app.post("/v1/chat/completions")
async def chat_completions(request: Request):
    """核心转发接口"""
    if not API_KEY:
        return JSONResponse(
            status_code=500,
            content={"error": "API_KEY 未设置，请在环境变量中配置"},
        )
    
    body = await request.json()
    messages = body.get("messages", [])
    
    # ---------- 提取用户最新消息 ----------
    user_message = ""
    for msg in reversed(messages):
        if msg.get("role") == "user":
            content = msg.get("content", "")
            if isinstance(content, str):
                user_message = content
            elif isinstance(content, list):
                user_message = " ".join(
                    item.get("text", "") for item in content
                    if isinstance(item, dict) and item.get("type") == "text"
                )
            break
    
    # ---------- 构建 system prompt ----------
    # 内部请求（如压缩上下文）可跳过人设注入
    skip_prompt = body.pop('skip_system_prompt', False)
    
    # 读取前端传来的模板变量上下文
    template_ctx = {
        'model_name': body.get('model', ''),
        'user_name': body.pop('user_name', ''),
        'assistant_name': body.pop('assistant_name', ''),
    }
    
    # v5.8：项目 ID（前端传来，用于项目指令/记忆/文件注入）
    project_id = body.pop('project_id', None) or None

    mem_enabled = await get_memory_enabled()
    prompt_meta = {}
    if not skip_prompt:
        # v5.6：计算用户消息数（用于无缝切窗判断是第几轮）
        user_msg_count = sum(1 for m in messages if m.get('role') == 'user')
        if mem_enabled and user_message:
            enhanced_prompt, prompt_meta = await build_system_prompt_with_memories(user_message, user_msg_count=user_msg_count, project_id=project_id)
        else:
            # v5.4：即使记忆关闭，也从数据库优先读取 system prompt（降级到文件版本）
            enhanced_prompt = await get_active_system_prompt() or SYSTEM_PROMPT
        
        if enhanced_prompt:
            # 替换模板变量
            enhanced_prompt = replace_template_variables(enhanced_prompt, template_ctx)

            has_system = any(msg.get("role") == "system" for msg in messages)
            if has_system:
                for i, msg in enumerate(messages):
                    if msg.get("role") == "system":
                        messages[i]["content"] = enhanced_prompt + "\n\n" + msg["content"]
                        break
            else:
                messages.insert(0, {"role": "system", "content": enhanced_prompt})
            
            # ---- Prompt Caching：把 system message 拆成静态/动态两个 content block ----
            # 只对 Claude 模型做，其他模型自动缓存不需要
            _model_for_cache = body.get("model", "").lower()
            is_claude = "claude" in _model_for_cache or "anthropic" in _model_for_cache
            cache_enabled_val = await get_config("prompt_cache_enabled")
            cache_on = is_claude and (cache_enabled_val is None or str(cache_enabled_val).lower() != 'false')
            
            if cache_on:
                for i, msg in enumerate(messages):
                    if msg.get("role") == "system" and isinstance(msg.get("content"), str):
                        content = msg["content"]
                        if "<!-- CACHE_BOUNDARY -->" in content:
                            static_part, dynamic_part = content.split("<!-- CACHE_BOUNDARY -->", 1)
                            static_part = static_part.rstrip()
                            dynamic_part = dynamic_part.strip()
                            
                            content_blocks = [
                                {"type": "text", "text": static_part, "cache_control": {"type": "ephemeral"}}
                            ]
                            if dynamic_part:
                                content_blocks.append({"type": "text", "text": dynamic_part})
                            
                            messages[i]["content"] = content_blocks
                            print(f"💾 Prompt 缓存已启用：静态 ~{len(static_part)}字 + 动态 ~{len(dynamic_part)}字")
                        break
    
    # 替换前端传来的 skill prompt 中的模板变量
    for msg in messages:
        if msg.get("role") == "system" and '{' in msg.get("content", ""):
            msg["content"] = replace_template_variables(msg["content"], template_ctx)
    
    body["messages"] = messages
    
    # ---------- 联网搜索 ----------
    tool_events = []  # 收集工具事件，通过 SSE 发给前端展示

    web_search_mode = body.pop("web_search", False)
    # 兼容布尔和字符串：true → 强制搜索, "auto" → 模型自行决定, false → 关闭
    if web_search_mode == "auto":
        do_search_force = False
        do_search_auto = True
    else:
        do_search_force = bool(web_search_mode)
        do_search_auto = False

    if do_search_force and user_message:
        try:
            search_engine = await get_config("search_engine") or ""
            search_api_key = await get_config("search_api_key") or ""
            search_max = await get_config_int("search_max_results", fallback=5)
            
            if search_engine:
                print(f"🌐 联网搜索: [{search_engine}] {user_message[:60]}...")
                search_results = await web_search(
                    query=user_message[:200],
                    engine=search_engine,
                    api_key=search_api_key,
                    max_results=search_max,
                )
                if search_results:
                    search_text = format_results_for_prompt(search_results, user_message[:100])
                    messages.append({"role": "system", "content": search_text})
                    body["messages"] = messages
                    print(f"🌐 搜索完成，获得 {len(search_results)} 条结果")
                    tool_events.append({
                        "type": "search", "engine": search_engine,
                        "query": user_message[:100], "count": len(search_results),
                        "results": [r.to_dict() for r in search_results[:10]],
                    })
                else:
                    print(f"🌐 搜索无结果")
            else:
                print(f"⚠️ 联网搜索已请求但未配置搜索引擎")
        except Exception as e:
            print(f"❌ 联网搜索出错: {e}")
    
    # ---------- 模型处理 ----------
    model = body.get("model", DEFAULT_MODEL)
    if not model:
        model = DEFAULT_MODEL
    body["model"] = model
    
    # ---------- 供应商路由 ----------
    # 根据 model_id 查找已配置的供应商，找不到就用全局环境变量
    # 如果数据库不可用（DATABASE_URL 未设置），也降级到环境变量
    try:
        provider_info = await resolve_provider_for_model(model)
    except Exception:
        provider_info = None
    if provider_info:
        chat_api_key = provider_info["api_key"]
        # 确保 URL 以 /chat/completions 结尾
        base = provider_info["api_base_url"].rstrip("/")
        chat_api_url = base if base.endswith("/chat/completions") else f"{base}/chat/completions"
        print(f"🔀 路由到供应商 [{provider_info['provider_name']}]: {base}")
    else:
        chat_api_key = API_KEY
        chat_api_url = API_BASE_URL
    
    # ---------- MCP 工具调用 ----------
    mcp_servers = body.pop("mcp_servers", [])  # 从 body 中取出并移除
    
    # ---------- 生成 session ID ----------
    session_id = str(uuid.uuid4())[:8]
    
    # 请求 LLM 在流式响应中包含 token 用量
    if body.get("stream"):
        body.setdefault("stream_options", {})["include_usage"] = True
    
    # OpenRouter：配置思考链参数
    if "openrouter" in chat_api_url:
        reasoning_effort = body.pop("reasoning_effort", None)
        reasoning_cfg = {"enabled": True}
        if reasoning_effort and reasoning_effort in ("low", "medium", "high"):
            reasoning_cfg["effort"] = reasoning_effort
        body["reasoning"] = reasoning_cfg
    else:
        # 非 OpenRouter 供应商，reasoning_effort 保持原样传给 API（DeepSeek 等会忽略不认识的参数）
        pass
    
    # ---------- Prompt 缓存（v5.5 → v5.7 修正）----------
    # cache_control 现在在 system message 的 content block 上加，不在 body 顶层
    # （旧代码在 body 加 cache_control 是无效的，OpenRouter 需要 content block 级标记）
    
    # ---------- Claude Provider 偏好 ----------
    # 优先走 Anthropic 直连（缓存支持最好），允许回退
    model_lower_for_provider = model.lower()
    if ("claude" in model_lower_for_provider or "anthropic" in model_lower_for_provider) and "openrouter" in chat_api_url:
        if "provider" not in body:
            body["provider"] = {"order": ["Anthropic"], "allow_fallbacks": True}
            print(f"🔀 Provider 偏好：优先 Anthropic 直连")

    # ---------- 转发请求 ----------
    headers = {
        "Authorization": f"Bearer {chat_api_key}",
        "Content-Type": "application/json",
    }
    if "openrouter" in chat_api_url:
        headers["HTTP-Referer"] = EXTRA_REFERER
        headers["X-Title"] = EXTRA_TITLE
    
    is_stream = body.get("stream", False)
    
    # 🔍 调试：记录思考链相关参数
    if body.get("reasoning") or body.get("reasoning_effort") or body.get("include_reasoning"):
        print(f"🔍 [思考链参数] reasoning={body.get('reasoning')}, reasoning_effort={body.get('reasoning_effort')}, include_reasoning={body.get('include_reasoning')}")
    
    # ========== 收集工具（MCP + 联网搜索 auto） ==========
    openai_tools = []
    tool_map = {}

    # MCP 工具
    if mcp_servers:
        try:
            mcp_tools, mcp_map = await get_tools_for_servers(mcp_servers)
            openai_tools.extend(mcp_tools)
            tool_map.update(mcp_map)
        except Exception as e:
            print(f"❌ MCP 工具获取失败: {e}")

    # 联网搜索 auto 模式：注册为 function tool，让模型自行决定是否调用
    if do_search_auto:
        search_engine = await get_config("search_engine") or ""
        if search_engine:
            openai_tools.append({
                "type": "function",
                "function": {
                    "name": "_gateway_web_search",
                    "description": "搜索互联网获取实时信息。仅在用户明确要求联网搜索、或需要最新新闻/天气/实时数据/你不确定的事实时调用。闲聊、角色扮演、创意写作等不需要调用。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "query": {
                                "type": "string",
                                "description": "搜索关键词，用简洁的搜索引擎友好格式",
                            }
                        },
                        "required": ["query"],
                    },
                },
            })
            # 标记为网关内置工具（不走 MCP，本地执行）
            tool_map["_gateway_web_search"] = {"type": "gateway_builtin", "handler": "web_search"}
            print(f"🌐 联网搜索已注册为工具（auto 模式，引擎: {search_engine}）")
        else:
            print(f"⚠️ 联网搜索 auto 模式已请求但未配置搜索引擎")

        # v5.8: 对话搜索工具（关键词触发）
    _SEARCH_KEYWORDS = ["之前", "上次", "以前", "上回", "聊过", "说过", "提过", "提到过", "讨论过", "记得我们"]
    if mem_enabled and any(kw in user_message for kw in _SEARCH_KEYWORDS):
        openai_tools.append({
            "type": "function",
            "function": {
                "name": "_gateway_search_conversations",
                "description": "搜索过去的对话记录。当用户提到'我们之前聊过''上次说的''之前讨论的'等回忆性表达，或者你需要查找过去对话中的具体细节时调用。输入搜索关键词，返回匹配的对话片段和上下文。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {
                            "type": "string",
                            "description": "搜索关键词，用简洁的内容关键词（如'用药方案''生日''项目部署'），不要用'我们讨论过'之类的元描述",
                        },
                        "limit": {
                            "type": "integer",
                            "description": "最多返回几条匹配（默认10）",
                        },
                    },
                    "required": ["query"],
                },
            },
        })
        tool_map["_gateway_search_conversations"] = {"type": "gateway_builtin", "handler": "search_conversations", "project_id": project_id}
        print(f"🔍 对话搜索工具已注册")

    # 提醒系统工具：仅在消息可能涉及提醒时注册（省 API 调用）
    _REMINDER_TRIGGER_KEYWORDS = [
        # 创建提醒
        "提醒", "闹钟", "定时", "叫我", "别忘了", "不要忘", "记得提醒",
        "到时候", "之后叫", "之后提醒", "点钟", "点半",
        "每天", "每周", "每小时", "每隔",
        # 查看/管理提醒
        "有什么提醒", "哪些提醒", "设了什么", "取消提醒", "删除提醒",
        "不用提醒", "提醒列表", "做完了", "回来了", "学完了",
    ]
    _need_reminder_tools = any(kw in user_message for kw in _REMINDER_TRIGGER_KEYWORDS)

    if _need_reminder_tools:
        _reminder_tools = [
        {
            "type": "function",
            "function": {
                "name": "_gateway_create_reminder",
                "description": "为用户创建一条提醒。当用户说'提醒我...'、'...之后叫我...'、'每天...点提醒我...'时调用。title 用简洁的中文描述，notes 用来记录上下文信息以便提醒时参考。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title": {"type": "string", "description": "提醒标题，简短描述（如'吃药''给妈妈打电话'）"},
                        "notes": {"type": "string", "description": "备注信息，提醒时作为上下文参考（如'妈妈上周说周末要搬东西'）"},
                        "trigger_time": {"type": "string", "description": "触发时间，ISO 8601 格式（如'2026-03-31T23:00:00+08:00'）。相对时间请转换为绝对时间。"},
                        "repeat_type": {"type": "string", "enum": ["once", "daily", "weekly", "hourly"], "description": "重复类型：once=一次性, daily=每天, weekly=每周, hourly=每N小时"},
                        "repeat_config": {"type": "object", "description": "循环配置（hourly时传{hours:N}）"},
                    },
                    "required": ["title", "trigger_time"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "_gateway_list_reminders",
                "description": "查看用户当前的所有活跃提醒。当用户问'我设了哪些提醒'、'有什么提醒'时调用。",
                "parameters": {"type": "object", "properties": {}},
            },
        },
        {
            "type": "function",
            "function": {
                "name": "_gateway_complete_reminder",
                "description": "标记一条提醒为已完成。当用户表示事情已经做完（如'回来了''做完了''学完了'），且当前有相关的待触发提醒时调用。一次性提醒会被标记完成，循环提醒不受影响。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "reminder_id": {"type": "string", "description": "要完成的提醒 ID"},
                    },
                    "required": ["reminder_id"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "_gateway_delete_reminder",
                "description": "删除一条提醒（包括循环提醒）。当用户说'取消那个提醒'、'以后不用提醒我...了'时调用。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "reminder_id": {"type": "string", "description": "要删除的提醒 ID"},
                    },
                    "required": ["reminder_id"],
                },
            },
        },
    ]
        openai_tools.extend(_reminder_tools)
        for t in _reminder_tools:
            tool_map[t["function"]["name"]] = {"type": "gateway_builtin", "handler": "reminder"}
        print(f"⏰ 提醒工具已注册（关键词命中：{user_message[:30]}）")

    # ========== Tool Call 模式（MCP 和/或 auto 搜索） ==========
    if openai_tools and is_stream:
        print(f"🔧 工具模式: 共 {len(openai_tools)} 个工具可用")

        return StreamingResponse(
            _stream_with_tools(
                messages=messages,
                tools=openai_tools,
                tool_map=tool_map,
                model=model,
                temperature=body.get("temperature", 0.7),
                tool_events=tool_events,
                session_id=session_id,
                user_message=user_message,
                mem_enabled=mem_enabled,
                api_url=chat_api_url,
                api_key=chat_api_key,
                project_id=project_id,
                prompt_meta=prompt_meta,
            ),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
        )

    # ========== 正常转发模式 ==========
    if is_stream:
        return StreamingResponse(
            stream_and_capture(headers, body, session_id, user_message, model, tool_events, api_url=chat_api_url, project_id=project_id, prompt_meta=prompt_meta),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
        )
    else:
        async with httpx.AsyncClient(timeout=300) as client:
            response = await client.post(chat_api_url, headers=headers, json=body)
            
            if response.status_code == 200:
                resp_data = response.json()
                assistant_msg = ""
                try:
                    assistant_msg = resp_data["choices"][0]["message"]["content"]
                except (KeyError, IndexError):
                    pass
                
                if mem_enabled and user_message and assistant_msg:
                    _emo = merge_emotion_levels(detect_emotion_from_user_msg(user_message), detect_emotion_from_response(assistant_msg))
                    _spawn_background_task(
                        process_memories_background(session_id, user_message, assistant_msg, model, emotion_level=_emo, project_id=project_id)
                    )
                
                return JSONResponse(status_code=200, content=resp_data)
            else:
                try:
                    err_content = response.json()
                except Exception:
                    err_content = {"error": response.text[:500]}
                return JSONResponse(status_code=response.status_code, content=err_content)


async def _execute_gateway_tool(tool_name: str, arguments: dict, tool_info: dict) -> tuple:
    """
    执行网关内置工具（联网搜索、提醒系统等）。
    返回 (result_text, extra_metadata) 元组，extra_metadata 用于 SSE 事件附加信息。
    """
    extra = {}

    if tool_name == "_gateway_web_search":
        query = arguments.get("query", "")
        if not query:
            return "搜索关键词为空", extra
        try:
            search_engine = await get_config("search_engine") or ""
            search_api_key = await get_config("search_api_key") or ""
            search_max = await get_config_int("search_max_results", fallback=5)

            print(f"🌐 [auto] 模型请求联网搜索: [{search_engine}] {query[:80]}")
            results = await web_search(
                query=query[:200],
                engine=search_engine,
                api_key=search_api_key,
                max_results=search_max,
            )
            if results:
                extra = {
                    "engine": search_engine,
                    "query": query[:100],
                    "count": len(results),
                    "results": [r.to_dict() for r in results[:10]],
                }
                text = format_results_for_prompt(results, query[:100])
                print(f"🌐 [auto] 搜索完成，{len(results)} 条结果")
                return text, extra
            else:
                print(f"🌐 [auto] 搜索无结果")
                return f"搜索「{query}」无结果。", extra
        except Exception as e:
            print(f"❌ [auto] 搜索出错: {e}")
            return f"搜索失败: {e}", extra

    # ── v5.8：对话搜索工具 ──

    if tool_name in ("_gateway_search_conversations", "gateway_search_conversations"):
        query = arguments.get("query", "")
        if not query:
            return "搜索关键词为空", extra
        try:
            from database import search_chat_messages
            search_limit = arguments.get("limit", 10)
            search_project_id = tool_info.get("project_id")
            results = await search_chat_messages(query, project_id=search_project_id, limit=search_limit, context_size=2)
            
            title_matches = results.get("title_matches", [])
            msg_matches = results.get("message_matches", [])
            
            if not title_matches and not msg_matches:
                print(f"🔍 对话搜索 '{query}' → 无结果")
                return f"在过去的对话中没有找到与「{query}」相关的内容。", extra
            
            # 格式化结果给模型
            lines = []
            
            if title_matches:
                lines.append(f"## 标题匹配（{len(title_matches)} 个对话）")
                for t in title_matches:
                    dt = t.get("date", "")[:10] if t.get("date") else ""
                    lines.append(f"- [{t['title']}]（{dt}）")
            
            if msg_matches:
                lines.append(f"\n## 消息内容匹配（{len(msg_matches)} 个对话）")
                for conv in msg_matches:
                    dt = conv.get("date", "")[:10] if conv.get("date") else ""
                    lines.append(f"\n### {conv['title']}（{dt}）")
                    for match in conv["matches"][:3]:
                        for msg in match.get("context", []):
                            role = "用户" if msg["role"] == "user" else "助手"
                            marker = "→ " if msg.get("is_match") else "  "
                            content = msg["content"][:200]
                            lines.append(f"{marker}{role}: {content}")
                        lines.append("")
            
            text = "\n".join(lines)
            total = len(title_matches) + sum(len(c.get("matches", [])) for c in msg_matches)
            print(f"🔍 对话搜索 '{query}' → 标题{len(title_matches)}条 + 消息{len(msg_matches)}个对话")
            
            extra = {
                "query": query[:100],
                "title_count": len(title_matches),
                "message_count": len(msg_matches),
            }
            return text, extra
        except Exception as e:
            print(f"❌ 对话搜索出错: {e}")
            return f"搜索失败: {e}", extra

    # ── 提醒系统工具 ──

    if tool_name == "_gateway_create_reminder":
        try:
            title = arguments.get("title", "")
            if not title:
                return "提醒标题不能为空", extra
            reminder_data = {
                "title": title,
                "notes": arguments.get("notes", ""),
                "trigger_time": arguments.get("trigger_time", ""),
                "repeat_type": arguments.get("repeat_type", "once"),
                "repeat_config": arguments.get("repeat_config", {}),
            }
            result = await create_reminder(reminder_data)
            repeat_label = {"once": "一次性", "daily": "每天", "weekly": "每周", "hourly": "循环"}.get(reminder_data["repeat_type"], "一次性")
            print(f"⏰ 提醒已创建: [{repeat_label}] {title}")
            return json.dumps({
                "success": True,
                "message": f"提醒已创建：{title}（{repeat_label}）",
                "reminder": result,
            }, ensure_ascii=False), extra
        except Exception as e:
            print(f"❌ 创建提醒失败: {e}")
            return f"创建提醒失败: {e}", extra

    if tool_name == "_gateway_list_reminders":
        try:
            reminders = await get_reminders(include_completed=False)
            if not reminders:
                return json.dumps({"success": True, "message": "当前没有活跃的提醒", "reminders": []}, ensure_ascii=False), extra
            lines = []
            for r in reminders:
                repeat_label = {"once": "一次性", "daily": "每天", "weekly": "每周", "hourly": "循环"}.get(r.get("repeat_type", "once"), "")
                status = "✅" if r.get("enabled") else "⏸️"
                lines.append(f"{status} [{r['id']}] {r['title']}（{repeat_label}，{r['trigger_time']}）")
                if r.get("notes"):
                    lines.append(f"   备注: {r['notes']}")
            return json.dumps({
                "success": True,
                "message": f"共 {len(reminders)} 条活跃提醒",
                "details": "\n".join(lines),
                "reminders": reminders,
            }, ensure_ascii=False), extra
        except Exception as e:
            return f"查询提醒失败: {e}", extra

    if tool_name == "_gateway_complete_reminder":
        try:
            rid = arguments.get("reminder_id", "")
            if not rid:
                return "请提供提醒 ID", extra
            # 查找提醒信息，循环提醒用 fire_reminder 计算下次触发时间
            all_rems = await get_reminders(include_completed=False)
            rem = next((r for r in all_rems if r["id"] == rid), None)
            if not rem:
                return json.dumps({"success": False, "message": "提醒不存在或已完成"}, ensure_ascii=False), extra
            ok = await fire_reminder(rid, rem.get("repeat_type", "once"), rem.get("repeat_config"))
            if ok:
                action = "已标记为完成" if rem.get("repeat_type") == "once" else "已触发，下次将自动提醒"
                print(f"⏰ 提醒已处理: {rid} ({action})")
                return json.dumps({"success": True, "message": f"提醒「{rem.get('title', rid)}」{action}"}, ensure_ascii=False), extra
            return json.dumps({"success": False, "message": "操作失败"}, ensure_ascii=False), extra
        except Exception as e:
            return f"完成提醒失败: {e}", extra

    if tool_name == "_gateway_delete_reminder":
        try:
            rid = arguments.get("reminder_id", "")
            if not rid:
                return "请提供提醒 ID", extra
            ok = await delete_reminder(rid)
            if ok:
                print(f"⏰ 提醒已删除: {rid}")
                return json.dumps({"success": True, "message": f"提醒已删除"}, ensure_ascii=False), extra
            return json.dumps({"success": False, "message": "提醒不存在"}, ensure_ascii=False), extra
        except Exception as e:
            return f"删除提醒失败: {e}", extra

    return f"未知的内置工具: {tool_name}", extra


async def _stream_with_tools(messages, tools, tool_map, model, temperature, tool_events, session_id, user_message, mem_enabled, api_url=None, api_key=None, project_id=None, prompt_meta=None):
    """
    工具 + 流式模式：tool call 轮次用非流式（需要完整看 tool_calls），
    最终回复直接输出已获得的内容（模拟流式），不再重复请求 LLM。
    工具执行采用并发策略：同服务器复用连接，跨服务器并行。
    """
    import httpx as _httpx

    _api_url = api_url or API_BASE_URL
    _api_key = api_key or API_KEY

    # 先发送衔接提示（如果有无缝切窗）
    if prompt_meta and prompt_meta.get("handoff"):
        yield f"data: {json.dumps({'ev_handoff': prompt_meta['handoff']}, ensure_ascii=False)}\n\n"

    # 先发送已有的 tool_events（比如强制搜索结果）
    for evt in (tool_events or []):
        yield f"data: {json.dumps({'ev_tool': evt}, ensure_ascii=False)}\n\n"

    headers = {
        "Authorization": f"Bearer {_api_key}",
        "Content-Type": "application/json",
    }
    if "openrouter" in _api_url:
        headers["HTTP-Referer"] = EXTRA_REFERER
        headers["X-Title"] = EXTRA_TITLE

    current_messages = list(messages)
    max_rounds = 10

    for round_num in range(max_rounds):
        # ── tool call 轮：非流式请求，检测是否有工具调用 ──
        body = {
            "model": model,
            "messages": current_messages,
            "tools": tools,
            "tool_choice": "auto",
            "temperature": temperature,
            "stream": False,
        }
        # OpenRouter：非流式也启用思考链，这样最终回复直接输出时不丢思考内容
        if "openrouter" in _api_url:
            body["reasoning"] = {"enabled": True}

        print(f"🔄 Tool loop round {round_num + 1}: {len(tools)} tools, {len(current_messages)} msgs")

        async with _httpx.AsyncClient(timeout=120) as client:
            resp = await client.post(_api_url, headers=headers, json=body)
            if resp.status_code != 200:
                print(f"❌ LLM 请求失败: {resp.status_code}")
                yield f"data: {json.dumps({'choices': [{'delta': {'content': f'⚠️ 模型请求失败 ({resp.status_code})'}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)}\n\n"
                yield "data: [DONE]\n\n"
                return
            data = resp.json()

        choice = data.get("choices", [{}])[0]
        message = choice.get("message", {})
        tool_calls = message.get("tool_calls", [])

        print(f"🔄 Round {round_num + 1}: tool_calls={len(tool_calls)}, has_content={bool(message.get('content'))}")

        if not tool_calls:
            # ── 无 tool_calls：直接用非流式结果，模拟流式输出 ──
            # v5.4 优化：不再重发流式请求，省掉一次完整的模型调用延迟
            final_text = message.get("content", "")
            usage_data = data.get("usage")

            if round_num == 0:
                print(f"⚡ 第一轮无工具调用，直接复用结果输出（省去二次请求）")
            else:
                print(f"✅ 工具调用后最终回复：直接输出 {len(final_text)} 字符")

            # 处理思考链
            reasoning = message.get("reasoning_content") or message.get("reasoning") or ""
            if reasoning and isinstance(reasoning, str):
                for i in range(0, len(reasoning), 40):
                    yield f"data: {json.dumps({'choices': [{'delta': {'reasoning_content': reasoning[i:i+40]}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)}\n\n"

            # 模拟流式输出正文
            if final_text:
                chunk_size = 20
                for i in range(0, len(final_text), chunk_size):
                    yield f"data: {json.dumps({'choices': [{'delta': {'content': final_text[i:i+chunk_size]}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)}\n\n"
                    await asyncio.sleep(0.008)

            finish_payload = {'choices': [{'delta': {}, 'finish_reason': 'stop'}], 'model': model}
            if usage_data:
                finish_payload['usage'] = usage_data
            yield f"data: {json.dumps(finish_payload, ensure_ascii=False)}\n\n"
            yield "data: [DONE]\n\n"

            assistant_msg = final_text
            if mem_enabled and user_message and assistant_msg:
                _emo = merge_emotion_levels(detect_emotion_from_user_msg(user_message), detect_emotion_from_response(assistant_msg))
                mem_result = await process_memories_background(session_id, user_message, assistant_msg, model, emotion_level=_emo, project_id=project_id)
                if mem_result and mem_result.get("action") != "skip":
                    yield f"data: {json.dumps({'ev_memory': mem_result}, ensure_ascii=False)}\n\n"
            return

        # ── 有 tool_calls → 并行执行工具 ──
        current_messages.append({
            "role": "assistant",
            "content": message.get("content") or None,
            "tool_calls": tool_calls,
        })

        # 解析所有工具调用
        parsed = []
        for tc in tool_calls:
            tc_id = tc.get("id", "")
            func = tc.get("function", {})
            tc_name = func.get("name", "")
            tc_args_str = func.get("arguments", "{}")
            try:
                tc_args = json.loads(tc_args_str)
            except json.JSONDecodeError:
                tc_args = {}
            parsed.append({"id": tc_id, "name": tc_name, "args": tc_args})

        # 分组：网关内置工具 vs MCP 远程工具
        # v5.8：兼容模型可能吃掉工具名前缀 _ 的情况
        def _resolve_tool_name(name):
            if name in tool_map:
                return name
            if f"_{name}" in tool_map:
                return f"_{name}"
            return name
        
        for p in parsed:
            p["name"] = _resolve_tool_name(p["name"])
        
        gw_parsed = [p for p in parsed if tool_map.get(p["name"], {}).get("type") == "gateway_builtin"]
        mcp_parsed = [p for p in parsed if tool_map.get(p["name"], {}).get("type") != "gateway_builtin"]

        tool_results = {}   # { call_id: result_text }
        tool_extras = {}    # { call_id: extra_metadata } — 并发安全，每个 call 独立

        # 网关工具（联网搜索等）：各自并发
        async def _run_gw(p):
            tool_info = tool_map.get(p["name"], {})
            result_text, extra_meta = await _execute_gateway_tool(p["name"], p["args"], tool_info)
            tool_results[p["id"]] = result_text
            tool_extras[p["id"]] = extra_meta

        # 构建并发任务列表
        tasks = [_run_gw(p) for p in gw_parsed]

        # MCP 工具：同服务器复用连接，不同服务器并发
        if mcp_parsed:
            async def _run_mcp():
                r = await call_tools_batch(mcp_parsed, tool_map)
                tool_results.update(r)
            tasks.append(_run_mcp())

        # 所有工具并发执行
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)

        print(f"⚡ {len(parsed)} 个工具调用并发完成")

        # 发送工具事件给前端 + 加入消息历史
        for p in parsed:
            result_text = tool_results.get(p["id"], "执行失败")

            evt_type = "search" if p["name"] in ("_gateway_web_search", "gateway_web_search") else "tool_call"
            evt = {
                "type": evt_type, "name": p["name"],
                "arguments": p["args"],
                "result": result_text[:2000] if result_text else "",
            }
            sr = tool_extras.get(p["id"], {})
            if sr:
                evt.update(sr)
            yield f"data: {json.dumps({'ev_tool': evt}, ensure_ascii=False)}\n\n"

            current_messages.append({
                "role": "tool",
                "tool_call_id": p["id"],
                "content": result_text[:8000] if result_text else "",
            })

    # 循环结束还没出结果
    yield f"data: {json.dumps({'choices': [{'delta': {'content': '⚠️ 工具调用轮次过多，已停止'}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)}\n\n"
    yield "data: [DONE]\n\n"


async def _simulate_stream(text: str, model: str, tool_events: list = None):
    """将完整文本模拟为 SSE 流式输出（tool call 完成后使用）"""
    # 先发送工具事件
    for evt in (tool_events or []):
        yield f"data: {json.dumps({'ev_tool': evt}, ensure_ascii=False)}\n\n"
    
    chunk_size = 20
    for i in range(0, len(text), chunk_size):
        chunk = text[i:i + chunk_size]
        data = {
            "choices": [{
                "delta": {"content": chunk},
                "finish_reason": None,
            }],
            "model": model,
        }
        yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"
        await asyncio.sleep(0.02)  # 模拟流式延迟
    
    # 发送结束标记
    yield "data: [DONE]\n\n"


def _estimate_tokens(text: str) -> int:
    """估算文本的 token 数（中文 ~1.5 char/token，英文 ~4 char/token）"""
    if not text:
        return 0
    cjk = sum(1 for ch in text if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef')
    other = len(text) - cjk
    return max(1, round(cjk / 1.5 + other / 4))


async def stream_and_capture(headers: dict, body: dict, session_id: str, user_message: str, model: str, tool_events: list = None, api_url: str = None, project_id: str = None, prompt_meta: dict = None):
    """流式响应 + 捕获完整回复 + 工具事件"""
    _api_url = api_url or API_BASE_URL
    
    # 先发送衔接提示（如果有无缝切窗）
    if prompt_meta and prompt_meta.get("handoff"):
        yield f"data: {json.dumps({'ev_handoff': prompt_meta['handoff']}, ensure_ascii=False)}\n\n".encode("utf-8")
    
    # 先发送工具事件
    for evt in (tool_events or []):
        yield f"data: {json.dumps({'ev_tool': evt}, ensure_ascii=False)}\n\n".encode("utf-8")
    
    full_response = []
    buffer = ""
    _logged_first_delta = False
    _reasoning_chunks = 0
    
    async with httpx.AsyncClient(timeout=300) as client:
        async with client.stream("POST", _api_url, headers=headers, json=body) as response:
            if response.status_code != 200:
                error_body = b""
                async for chunk in response.aiter_bytes():
                    error_body += chunk
                print(f"❌ 流式请求失败 [{response.status_code}]: {error_body[:500].decode('utf-8', errors='ignore')}")
                err_msg = f"⚠️ 请求失败 ({response.status_code})"
                err_payload = json.dumps({'choices': [{'delta': {'content': err_msg}, 'finish_reason': None}], 'model': model}, ensure_ascii=False)
                yield f"data: {err_payload}\n\n".encode("utf-8")
                yield b"data: [DONE]\n\n"
                return
            
            async for chunk in response.aiter_bytes():
                yield chunk
                try:
                    buffer += chunk.decode("utf-8", errors="ignore")
                    while "\n" in buffer:
                        line, buffer = buffer.split("\n", 1)
                        line = line.strip()
                        if line.startswith("data: ") and line != "data: [DONE]":
                            try:
                                data = json.loads(line[6:])
                                delta = data.get("choices", [{}])[0].get("delta", {})
                                
                                # 🔍 调试日志：记录第一个有效delta的所有字段
                                if not _logged_first_delta and delta:
                                    keys = list(delta.keys())
                                    if keys and keys != ['role']:
                                        print(f"🔍 [流式调试] 首个delta字段: {keys}, 模型: {model}")
                                        # 如果有思考链相关字段，打印示例
                                        for k in ('reasoning_content', 'reasoning', 'reasoning_details'):
                                            if k in delta:
                                                sample = str(delta[k])[:100]
                                                print(f"🔍 [流式调试] {k} 示例: {sample}")
                                        _logged_first_delta = True
                                
                                # 统计思考链 chunk 数
                                if delta.get('reasoning_content') or delta.get('reasoning') or delta.get('reasoning_details'):
                                    _reasoning_chunks += 1
                                
                                content = delta.get("content", "")
                                if content:
                                    full_response.append(content)
                            except (json.JSONDecodeError, KeyError, IndexError):
                                pass
                except Exception:
                    pass
    
    assistant_msg = "".join(full_response)
    
    # 🔍 流式完成汇总
    print(f"🔍 [流式完成] 模型={model}, 正文={len(assistant_msg)}字, 思考链chunks={_reasoning_chunks}")
    if _reasoning_chunks == 0 and '<think>' in assistant_msg:
        print(f"🔍 [流式完成] ⚠️ 思考链在正文中（<think>标签），前端需要解析")
    
    # 🩷 情绪检测（v5.2）
    user_emotion = detect_emotion_from_user_msg(user_message)
    response_emotion = detect_emotion_from_response(assistant_msg)
    emotion_level = merge_emotion_levels(user_emotion, response_emotion)
    if emotion_level != "normal":
        print(f"🩷 情绪检测: user={user_emotion}, response={response_emotion} → {emotion_level}")
    
    if await get_memory_enabled() and user_message and assistant_msg:
        mem_result = await process_memories_background(session_id, user_message, assistant_msg, model, emotion_level=emotion_level, project_id=project_id)
        if mem_result and mem_result.get("action") != "skip":
            yield f"data: {json.dumps({'ev_memory': mem_result}, ensure_ascii=False)}\n\n".encode("utf-8")


# ============================================================
# 记忆管理接口
# ============================================================

@app.get("/debug/memories")
async def debug_memories(
    q: str = "",
    limit: int = 20,
    offset: int = 0,
    sort: str = "newest",
    category_id: int = None,
    min_importance: int = None,
):
    """查看和搜索记忆（支持分类筛选、分页、排序、重要度过滤）。

    返回字段名为 memories（前端约定），输出每条包含 is_permanent 和 heat
    （前端用于显示锁定 🔒 和热度 🔥 badge）。
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用（设置 MEMORY_ENABLED=true 开启）"}

    # 限制查询范围，防止过大请求消耗资源
    limit = max(1, min(limit, 200))
    offset = max(0, offset)

    try:
        if q:
            # 搜索路径：search_memories 已经返回 is_permanent / heat
            memories = await search_memories(q, limit=limit + offset, track_recall=False)
            if category_id is not None:
                memories = [m for m in memories if m.get("category_id") == category_id]
            if min_importance is not None:
                memories = [m for m in memories if m.get("importance", 0) >= min_importance]
            memories = memories[offset:offset + limit]
        else:
            # 非搜索路径：get_recent_memories 不算 heat, 这里在端点层补上
            memories = await get_recent_memories(limit=limit + offset, category_id=category_id)
            memories = [dict(m) for m in memories]
            if min_importance is not None:
                memories = [m for m in memories if m.get("importance", 0) >= min_importance]
            # 排序
            if sort == "oldest":
                memories.sort(key=lambda m: m.get("created_at") or "")
            elif sort == "importance":
                memories.sort(key=lambda m: m.get("importance", 0), reverse=True)
            elif sort == "heat":
                # heat 此路径不算，按 access_count 近似
                memories.sort(key=lambda m: m.get("access_count", 0) or 0, reverse=True)
            memories = memories[offset:offset + limit]

        total = await get_all_memories_count()

        return {
            "total_memories": total,
            "query": q or "(最近记忆)",
            "memories": [
                {
                    "id": m["id"],
                    "title": m.get("title", ""),
                    "content": m["content"],
                    "importance": m["importance"],
                    "is_permanent": m.get("is_permanent", False),
                    "heat": m.get("heat", 0),
                    "created_at": str(m["created_at"]),
                    "memory_type": m.get("memory_type", "fragment"),
                    "category_id": m.get("category_id"),
                    "category_name": m.get("category_name", ""),
                    "category_color": m.get("category_color", ""),
                    "source": m.get("source", "ai_extracted"),
                    "resolution": m.get("resolution", 1.0),
                }
                for m in memories
            ],
        }
    except Exception as e:
        return {"error": str(e)}


@app.delete("/debug/memories/{memory_id}")
async def delete_single_memory(memory_id: int):
    """删除单条记忆"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    
    try:
        success = await delete_memory(memory_id)
        if success:
            total = await get_all_memories_count()
            return {"status": "deleted", "memory_id": memory_id, "remaining": total}
        else:
            return JSONResponse(status_code=404, content={"error": f"记忆 #{memory_id} 不存在"})
    except Exception as e:
        return {"error": str(e)}


@app.post("/debug/memories/batch-delete")
async def batch_delete_memories(request: Request):
    """批量删除记忆（一次请求，一条 SQL）"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        body = await request.json()
        ids = body.get("ids", [])
        if not ids:
            return {"error": "ids 不能为空"}
        pool = await get_pool()
        async with pool.acquire() as conn:
            result = await conn.execute(
                "DELETE FROM memories WHERE id = ANY($1::int[])", ids
            )
        try:
            deleted = int(result.split(" ")[-1]) if result else 0
        except (ValueError, IndexError):
            deleted = 0
        total = await get_all_memories_count()
        return {"status": "deleted", "deleted": deleted, "remaining": total}
    except Exception as e:
        return {"error": str(e)}


@app.post("/debug/memories/batch-update")
async def batch_update_memories(request: Request):
    """批量更新记忆字段（importance / category_id / is_permanent）"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        body = await request.json()
        ids = body.get("ids", [])
        if not ids:
            return {"error": "ids 不能为空"}
        
        importance = body.get("importance")
        category_id = body.get("category_id", "UNSET")
        is_permanent = body.get("is_permanent")
        
        pool = await get_pool()
        async with pool.acquire() as conn:
            # 构建动态 SET 子句
            sets = []
            vals = []
            idx = 1
            if importance is not None:
                sets.append(f"importance = ${idx}")
                vals.append(importance)
                idx += 1
            if category_id != "UNSET":
                sets.append(f"category_id = ${idx}")
                vals.append(category_id)
                idx += 1
            if is_permanent is not None:
                sets.append(f"is_permanent = ${idx}")
                vals.append(bool(is_permanent))
                idx += 1
            
            if not sets:
                return {"error": "没有提供更新字段"}
            
            vals.append(ids)
            sql = f"UPDATE memories SET {', '.join(sets)} WHERE id = ANY(${idx}::int[])"
            await conn.execute(sql, *vals)
        
        return {"status": "updated", "count": len(ids)}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/debug/memories")
async def clear_memories():
    """清空所有记忆"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    
    try:
        count = await clear_all_memories()
        return {"status": "cleared", "deleted_count": count}
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/memory-heat")
async def debug_memory_heat(limit: int = 50):
    """
    记忆热度报告（v5.2）
    查看每条记忆的热度、召回次数、情绪浓度、查询多样性
    """
    try:
        from database import get_memory_heat_report
        report = await get_memory_heat_report(limit=min(limit, 200))
        
        # 统计摘要
        if report:
            hot = sum(1 for r in report if r["heat"] > 0.7)
            warm = sum(1 for r in report if 0.3 < r["heat"] <= 0.7)
            cold = sum(1 for r in report if r["heat"] <= 0.3)
            total_recalls = sum(r["access_count"] for r in report)
            emotional = sum(1 for r in report if r["emotional_weight"] > 0)
        else:
            hot = warm = cold = total_recalls = emotional = 0
        
        return {
            "summary": {
                "total": len(report),
                "hot": hot,
                "warm": warm,
                "cold": cold,
                "total_recalls": total_recalls,
                "emotional_memories": emotional,
            },
            "memories": report,
        }
    except Exception as e:
        return {"error": str(e)}


@app.put("/debug/memories/{memory_id}")
async def update_single_memory(memory_id: int, request: Request):
    """
    更新单条记忆
    请求体示例：{"content": "新内容", "importance": 8}
    可以只传其中一个字段
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    
    try:
        body = await request.json()
        content = body.get("content")
        importance = body.get("importance")
        title = body.get("title")
        # category_id: None清除分类, int设置分类, 不传不改
        cat_id = body.get("category_id", "UNSET")
        
        success = await update_memory(memory_id, content=content, importance=importance, title=title, category_id=cat_id)
        if success:
            return {"status": "updated", "memory_id": memory_id}
        else:
            return JSONResponse(status_code=404, content={"error": f"记忆 #{memory_id} 不存在或没有提供更新内容"})
    except Exception as e:
        return {"error": str(e)}


@app.post("/debug/memories")
async def add_memory_manual(request: Request):
    """
    手动添加记忆
    请求体示例：{"content": "用户喜欢喝奶茶", "importance": 7}
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    
    try:
        body = await request.json()
        content = body.get("content", "")
        importance = body.get("importance", 5)
        title = body.get("title", "")
        category_id = body.get("category_id")
        
        if not content:
            return JSONResponse(status_code=400, content={"error": "content 不能为空"})
        
        await save_memory(content=content, importance=importance, source_session="manual", title=title, category_id=category_id, source="user_explicit")
        total = await get_all_memories_count()
        return {"status": "added", "content": content, "importance": importance, "title": title, "total": total}
    except Exception as e:
        return {"error": str(e)}


@app.post("/debug/memories/{memory_id}/toggle-permanent")
async def toggle_memory_permanent(memory_id: int):
    """切换记忆的锁定状态"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        pool = await get_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT id, COALESCE(is_permanent, false) as is_permanent FROM memories WHERE id = $1",
                memory_id
            )
            if not row:
                return JSONResponse(status_code=404, content={"error": f"记忆 #{memory_id} 不存在"})
            
            new_val = not row["is_permanent"]
            await conn.execute(
                "UPDATE memories SET is_permanent = $1 WHERE id = $2",
                new_val, memory_id
            )
            status = "locked" if new_val else "unlocked"
            print(f"🔒 记忆 #{memory_id} {'锁定' if new_val else '解锁'}")
            return {"status": status, "memory_id": memory_id, "is_permanent": new_val}
    except Exception as e:
        return {"error": str(e)}


@app.get("/import/seed-memories")
async def import_seed_memories():
    """一次性导入预置记忆（从 seed_memories.py）"""
    try:
        from seed_memories import run_seed_import
        result = await run_seed_import()
        return result
    except ImportError:
        return {"error": "未找到 seed_memories.py，请参考 seed_memories_example.py 创建"}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/import/seed-memories")
async def clear_seed_memories():
    """清除所有种子记忆（source_session = 'seed-import'）"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        pool = await get_pool()
        async with pool.acquire() as conn:
            result = await conn.execute(
                "DELETE FROM memories WHERE source_session = 'seed-import'"
            )
        try:
            deleted = int(result.split(" ")[-1]) if result else 0
        except (ValueError, IndexError):
            deleted = 0
        total = await get_all_memories_count()
        return {"status": "cleared", "deleted": deleted, "remaining": total}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/migrate-embeddings")
async def api_migrate_embeddings():
    """为所有缺少向量的记忆生成 embedding"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        result = await migrate_embeddings()
        return result
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/embedding-stats")
async def api_embedding_stats():
    """查看 embedding 覆盖率"""
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        stats = await get_embedding_stats()
        return stats
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/extract-now")
async def api_extract_now(request: Request):
    """
    手动触发记忆提取（从最近对话中提取记忆）
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        # 解析 project_id
        project_id = None
        try:
            body = await request.json()
            project_id = body.get("project_id")
        except Exception:
            pass
        # 获取最近的对话消息
        extract_interval = await get_extract_interval()
        recent_msgs = await get_recent_conversation(limit=extract_interval * 2)
        if not recent_msgs:
            return {"status": "ok", "action": "extract", "saved": 0, "skipped": 0, "message": "没有最近的对话可提取"}
        
        messages_for_extraction = [
            {"role": row["role"], "content": row["content"]}
            for row in recent_msgs
        ]
        
        # 获取对比用的已有记忆
        user_text = " ".join(r["content"] for r in recent_msgs if r["role"] == "user")
        related = await search_memories(user_text[:500], limit=50, track_recall=False, project_id=project_id)
        recent = await get_recent_memories(limit=30, project_id=project_id)
        seen = set()
        existing_contents = []
        for content in [r["content"] for r in related] + [r["content"] for r in recent]:
            if content not in seen:
                seen.add(content)
                existing_contents.append(content)
        
        # 获取分类
        try:
            all_cats = await get_all_categories()
            cat_names = [c["name"] for c in all_cats]
        except Exception:
            cat_names = []
        
        from config import get_config
        db_memory_model = await get_config("default_memory_model")
        db_memory_prompt = await get_config("prompt_memory_extract")
        
        new_memories = await extract_memories(
            messages_for_extraction,
            existing_memories=existing_contents,
            categories=cat_names,
            model_override=db_memory_model if db_memory_model else None,
            prompt_override=db_memory_prompt if db_memory_prompt else None,
        )
        
        # 过滤 + 去重 + 保存
        META_BLACKLIST = [
            "记忆库", "记忆系统", "检索", "没有被记录", "没有被提取",
            "记忆遗漏", "尚未被记录", "写入不完整", "检索功能",
            "系统没有返回", "关键词匹配", "语义匹配", "语义检索",
            "阈值", "数据库", "seed", "导入", "部署",
            "bug", "debug", "端口", "网关",
        ]
        saved_count = 0
        skipped_count = 0
        session_id = "manual-" + str(uuid.uuid4())[:8]
        
        for mem in new_memories:
            if any(kw in mem["content"] for kw in META_BLACKLIST):
                continue
            is_dup, _ = await check_memory_duplicate(mem["content"], new_title=mem.get("title", ""))
            if is_dup:
                skipped_count += 1
                continue
            cat_id = None
            cat_hint = mem.get("category", "")
            if cat_hint:
                cat_id = await match_category_by_name(cat_hint)
            await save_memory(
                content=mem["content"],
                importance=mem["importance"],
                source_session=session_id,
                title=mem.get("title", ""),
                category_id=cat_id,
                source="manual_extracted",
                emotional_weight=mem.get("emotional_weight", 0),
                project_id=project_id,
            )
            saved_count += 1
        
        total = await get_all_memories_count()
        return {"status": "ok", "action": "extract", "saved": saved_count, "skipped": skipped_count, "total": total}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/daily-digest")
async def api_daily_digest(date: str = None):
    """
    手动触发每日记忆整理
    ?date=2026-03-02  指定日期整理
    不传 date 则整理昨天的
    """
    if not await get_memory_enabled():
        return {"error": "记忆系统未启用"}
    try:
        from daily_digest import run_daily_digest
        from config import get_config
        db_digest_model = await get_config("default_digest_model")
        db_digest_prompt = await get_config("prompt_daily_digest")
        result = await run_daily_digest(
            target_date=date,
            model_override=db_digest_model if db_digest_model else None,
            prompt_override=db_digest_prompt if db_digest_prompt else None,
        )
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 日历记忆页面接口（v5.0 记忆桥）
# ============================================================

@app.get("/admin/day-page")
async def api_generate_day_page(date: str = None):
    """手动触发日页面生成 ?date=2026-04-01"""
    try:
        from daily_digest import generate_day_page
        result = await generate_day_page(target_date=date)
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/week-summary")
async def api_generate_week_summary(start: str = None, end: str = None):
    """手动触发周总结 ?start=2026-03-31&end=2026-04-06"""
    try:
        from daily_digest import generate_week_summary
        if not start or not end:
            from datetime import timedelta as td, timezone as tz_mod, datetime as dt_cls
            TZ = tz_mod(td(hours=8))
            now = dt_cls.now(TZ)
            # 默认上周一到上周日
            days_since_monday = now.weekday()
            last_monday = now - td(days=days_since_monday + 7)
            last_sunday = last_monday + td(days=6)
            start = last_monday.strftime("%Y-%m-%d")
            end = last_sunday.strftime("%Y-%m-%d")
        result = await generate_week_summary(start, end)
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/month-summary")
async def api_generate_month_summary(month: str = None):
    """手动触发月总结 ?month=2026-03"""
    try:
        from daily_digest import generate_month_summary
        if not month:
            from datetime import timedelta as td, timezone as tz_mod, datetime as dt_cls
            TZ = tz_mod(td(hours=8))
            now = dt_cls.now(TZ)
            last_month_end = now.replace(day=1) - td(days=1)
            month = last_month_end.strftime("%Y-%m")
        # 解析月份
        year, mon = month.split("-")
        import calendar as cal_mod
        last_day = cal_mod.monthrange(int(year), int(mon))[1]
        start = f"{month}-01"
        end = f"{month}-{last_day:02d}"
        result = await generate_month_summary(start, end, month)
        return {"status": "ok", **result}
    except Exception as e:
        return {"error": str(e)}


@app.get("/calendar/{date}")
async def api_get_calendar_day(date: str, type: str = "day"):
    """获取指定日期的日历页面"""
    try:
        from database import get_calendar_page
        page = await get_calendar_page(date, type)
        if not page:
            return {"status": "ok", "page": None}
        # 序列化 date 对象
        page["date"] = str(page["date"])
        if page.get("created_at"):
            page["created_at"] = page["created_at"].isoformat()
        if page.get("updated_at"):
            page["updated_at"] = page["updated_at"].isoformat()
        return {"status": "ok", "page": page}
    except Exception as e:
        return {"error": str(e)}


@app.get("/calendar")
async def api_get_calendar_range(start: str = None, end: str = None, type: str = None):
    """获取一段时间的日历页面 ?start=2026-03-25&end=2026-04-01&type=day"""
    try:
        from database import get_calendar_range
        if not start or not end:
            # 默认最近7天
            from datetime import timedelta as td, timezone as tz_mod
            TZ = tz_mod(td(hours=8))
            from datetime import datetime as dt_cls
            now = dt_cls.now(TZ)
            if not end:
                end = now.strftime("%Y-%m-%d")
            if not start:
                start = (now - td(days=7)).strftime("%Y-%m-%d")
        pages = await get_calendar_range(start, end, type)
        for p in pages:
            p["date"] = str(p["date"])
            if p.get("created_at"):
                p["created_at"] = p["created_at"].isoformat()
            if p.get("updated_at"):
                p["updated_at"] = p["updated_at"].isoformat()
        return {"status": "ok", "pages": pages, "count": len(pages)}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/calendar/{date}")
async def api_save_calendar_page(date: str, req: Request):
    """用户手动编辑/创建日历页面"""
    try:
        from database import save_calendar_page
        body = await req.json()
        content = body.get("content", "")
        title = body.get("title", "")
        page_type = body.get("type", "day")
        # 用户编辑的内容存入 diary 字段，sections 留空（用户不走分段逻辑）
        page_id = await save_calendar_page(
            date_str=date,
            page_type=page_type,
            sections=[],
            diary=content,
            keywords=[],
            model_used="user_edit",
            summary="",
            digest="",
            title=title,
        )
        return {"status": "ok", "id": page_id}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/calendar/{date}")
async def api_delete_calendar_page(date: str, type: str = "day"):
    """删除指定日期的日历页面"""
    try:
        from database import delete_calendar_page
        ok = await delete_calendar_page(date, type)
        return {"status": "ok" if ok else "not_found"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 评论接口（v5.0 记忆桥通用）
# ============================================================

@app.post("/comments")
async def api_create_comment(req: Request):
    """创建评论"""
    try:
        from database import create_comment
        body = await req.json()
        comment = await create_comment(
            target_type=body["target_type"],
            target_id=body["target_id"],
            content=body["content"],
            author=body.get("author", "user"),
            parent_id=body.get("parent_id"),
        )
        if comment and comment.get("created_at"):
            comment["created_at"] = comment["created_at"].isoformat()
        return {"status": "ok", "comment": comment}
    except Exception as e:
        return {"error": str(e)}


@app.get("/comments")
async def api_get_comments(target_type: str, target_id: int):
    """获取评论列表 ?target_type=day_page&target_id=1"""
    try:
        from database import get_comments
        comments = await get_comments(target_type, target_id)
        for c in comments:
            if c.get("created_at"):
                c["created_at"] = c["created_at"].isoformat()
        return {"status": "ok", "comments": comments}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/comments/{comment_id}")
async def api_delete_comment(comment_id: int):
    """删除评论"""
    try:
        from database import delete_comment
        ok = await delete_comment(comment_id)
        return {"status": "ok" if ok else "not_found"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# Dream 接口（v5.1）
# ============================================================

@app.post("/dream/start")
async def api_dream_start(req: Request):
    """
    触发 Dream，返回 SSE 事件流
    Body: {"trigger_type": "manual"} (可选)
    """
    from starlette.responses import StreamingResponse
    from dream import run_dream

    body = {}
    try:
        body = await req.json()
    except Exception:
        pass

    trigger = body.get("trigger_type", "manual")

    async def event_generator():
        async for event in run_dream(trigger_type=trigger):
            event_type = event.get("type", "message")
            data = event.get("data", "")
            if isinstance(data, dict):
                data = json.dumps(data, ensure_ascii=False)
            yield f"event: {event_type}\ndata: {data}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.post("/dream/stop")
async def api_dream_stop():
    """中断正在进行的 Dream"""
    from dream import stop_dream
    return await stop_dream()


@app.post("/admin/dream/force-stop")
async def api_dream_force_stop():
    """强制清理卡住的 Dream（直接更新数据库状态）"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        result = await conn.execute("""
            UPDATE dream_logs SET status = 'interrupted', finished_at = NOW(),
                dream_narrative = COALESCE(dream_narrative, '') || '\n[手动强制中断]'
            WHERE status = 'running'
        """)
    return {"status": "ok", "message": f"已强制中断所有 running 状态的 Dream", "result": str(result)}


@app.get("/dream/status")
async def api_dream_status():
    """获取当前 Dream 状态"""
    try:
        from database import get_dream_status, get_unprocessed_memories
        from config import get_config
        status = await get_dream_status()
        unprocessed = await get_unprocessed_memories()
        last_dream_date = await get_config("last_dream_date")
        drowsy_threshold = int(await get_config("dream_drowsy_threshold") or "30")

        # 序列化时间
        for key in ("current", "last_completed"):
            if status.get(key):
                for field in ("started_at", "finished_at"):
                    if status[key].get(field):
                        status[key][field] = status[key][field].isoformat()

        return {
            "status": "ok",
            **status,
            "unprocessed_count": len(unprocessed),
            # 前端 admin-panel 读 unprocessed_fragments / is_dreaming, 这里加别名保证两套字段都能用
            "unprocessed_fragments": len(unprocessed),
            "is_dreaming": status.get("is_running", False),
            "drowsy_threshold": drowsy_threshold,
            "is_drowsy": len(unprocessed) >= drowsy_threshold,
            "last_dream_date": last_dream_date,
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/dream/history")
async def api_dream_history(limit: int = 10):
    """获取 Dream 执行历史"""
    try:
        from database import get_dream_history
        history = await get_dream_history(limit)
        for h in history:
            for field in ("started_at", "finished_at"):
                if h.get(field):
                    h[field] = h[field].isoformat()
        return {"status": "ok", "history": history}
    except Exception as e:
        return {"error": str(e)}


@app.get("/dream/scenes")
async def api_get_scenes():
    """获取所有活跃的记忆场景"""
    try:
        from database import get_active_scenes
        scenes = await get_active_scenes()
        for s in scenes:
            for field in ("created_at", "updated_at"):
                if s.get(field):
                    s[field] = s[field].isoformat()
        return {"status": "ok", "scenes": scenes, "count": len(scenes)}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/dream/{dream_id}")
async def api_delete_dream(dream_id: int):
    """删除一条 Dream 日志及其关联的场景"""
    try:
        from database import get_pool
        pool = await get_pool()
        async with pool.acquire() as conn:
            # 先删关联的场景
            await conn.execute("DELETE FROM mem_scenes WHERE created_by_dream_id = $1", dream_id)
            # 再删 dream 日志本身
            result = await conn.execute("DELETE FROM dream_logs WHERE id = $1", dream_id)
        if "DELETE 0" in result:
            return {"error": f"Dream #{dream_id} 不存在"}
        return {"status": "ok", "deleted": dream_id}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/scene/{scene_id}")
async def api_update_scene(scene_id: int, req: Request):
    """用户手动编辑记忆场景（标题、叙事、远见）"""
    try:
        from database import update_mem_scene
        body = await req.json()
        kwargs = {}
        if "title" in body:
            kwargs["title"] = body["title"]
        if "narrative" in body:
            kwargs["narrative"] = body["narrative"]
        if "foresight" in body:
            kwargs["foresight"] = body["foresight"]
        if not kwargs:
            return {"error": "没有可更新的字段"}
        ok = await update_mem_scene(scene_id, **kwargs)
        return {"status": "ok" if ok else "not_found"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 动态配置管理接口（v3.1）
# ============================================================

@app.get("/admin/config")
async def api_get_config():
    """获取所有配置"""
    try:
        config = await get_all_config()
        return {"status": "ok", "config": config}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/config/{key}")
async def api_set_config(key: str, request: Request):
    """更新单个配置"""
    try:
        data = await request.json()
        value = str(data.get("value", ""))
        success = await set_config(key, value)
        if success:
            return {"status": "updated", "key": key, "value": value}
        else:
            return {"error": f"无效的配置项或值: {key}={value}"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# Prompt 出厂默认值管理（v5.6）
# ============================================================

HANDOFF_SUMMARY_PROMPT = """根据上一个对话的最后几条消息，生成一段简短的话题摘要，供新对话衔接用。

## 要求
- 用中文
- 概括用户最后在聊什么话题、情绪状态、未完成的事
- 200字以内
- 只输出摘要，不要前缀或解释

## 上一个对话的最后内容
{messages}"""


# ============================================================
# 无缝切窗摘要缓存与生成
# ============================================================

# 缓存：{ "conv_id": "xxx", "summary": "..." }
_handoff_summary_cache = {"conv_id": None, "summary": None}


async def _generate_handoff_summary(conv_id: str, handoff_msgs: list, prev_title: str):
    """后台生成切窗摘要并缓存，第 2 轮起使用"""
    global _handoff_summary_cache
    try:
        # 读取自定义 prompt（如果用户改过的话）
        prompt_template = await get_config("prompt_handoff_summary") or HANDOFF_SUMMARY_PROMPT
        
        # 拼消息文本
        lines = []
        for m in handoff_msgs:
            role_label = "用户" if m["role"] == "user" else "助手"
            content = m.get("content", "")
            if len(content) > 500:
                content = content[:500] + "…（截断）"
            lines.append(f"{role_label}: {content}")
        messages_text = "\n".join(lines)
        
        prompt = prompt_template.replace("{messages}", messages_text)
        
        # 确定模型：优先用配置的摘要专用模型，否则用通用后台模型
        use_model = await get_config("handoff_summary_model") or os.getenv("MEMORY_MODEL", "anthropic/claude-haiku-4")
        
        # 解析供应商端点
        try:
            from database import resolve_model_endpoint
            use_api_url, use_api_key = await resolve_model_endpoint(use_model)
        except Exception:
            use_api_url = os.getenv("MEMORY_API_BASE_URL", "") or os.getenv("API_BASE_URL", "https://openrouter.ai/api/v1/chat/completions")
            if not use_api_url.rstrip("/").endswith("/chat/completions"):
                use_api_url = f"{use_api_url.rstrip('/')}/chat/completions"
            use_api_key = os.getenv("MEMORY_API_KEY", "") or os.getenv("API_KEY", "")
        
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(
                use_api_url,
                headers={
                    "Authorization": f"Bearer {use_api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": use_model,
                    "max_tokens": 500,
                    "messages": [
                        {"role": "user", "content": prompt},
                    ],
                },
            )
            
            if response.status_code == 200:
                data = response.json()
                summary = data.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
                if summary:
                    _handoff_summary_cache = {"conv_id": conv_id, "summary": summary}
                    print(f"🔗 切窗摘要已生成并缓存（{len(summary)}字）：{summary[:80]}...")
                    return
        
        print(f"⚠️  切窗摘要生成失败: HTTP {response.status_code}")
    except Exception as e:
        print(f"⚠️  切窗摘要生成失败: {e}")


def _get_factory_prompts() -> dict:
    """收集所有出厂默认 prompt（从各模块常量中读取）"""
    from memory_extractor import EXTRACTION_PROMPT, EMOTION_HIGH_INSTRUCTION
    from daily_digest import (
        DIGEST_PROMPT, DEFAULT_PROFILE_PROMPT, DAY_PAGE_PROMPT,
        WEEK_SUMMARY_PROMPT, MONTH_SUMMARY_PROMPT, PERIOD_SUMMARY_PROMPT,
    )
    from dream import DREAM_PROMPT
    return {
        "prompt_memory_extract":    EXTRACTION_PROMPT,
        "prompt_daily_digest":      DIGEST_PROMPT,
        "prompt_user_profile":      DEFAULT_PROFILE_PROMPT,
        "prompt_daily_digest_page": DAY_PAGE_PROMPT,
        "prompt_weekly_summary":    WEEK_SUMMARY_PROMPT,
        "prompt_monthly_summary":   MONTH_SUMMARY_PROMPT,
        "prompt_period_summary":    PERIOD_SUMMARY_PROMPT,
        "prompt_dream":             DREAM_PROMPT,
        "prompt_handoff_summary":   HANDOFF_SUMMARY_PROMPT,
    }


@app.get("/admin/default-prompts")
async def api_get_default_prompts():
    """获取所有出厂默认 prompt（供前端「恢复默认」按钮使用）"""
    try:
        factory = _get_factory_prompts()
        return {"status": "ok", "prompts": factory}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/restore-prompt/{key}")
async def api_restore_prompt(key: str):
    """将指定 prompt 恢复为出厂默认值"""
    try:
        factory = _get_factory_prompts()
        if key not in factory:
            return {"error": f"未知的 prompt 配置项: {key}"}
        default_value = factory[key]
        success = await set_config(key, default_value)
        if success:
            print(f"🔄 已恢复默认 prompt: {key}")
            return {"status": "restored", "key": key, "length": len(default_value)}
        else:
            return {"error": f"写入失败: {key}"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 供应商管理 API
# ============================================================

@app.get("/admin/providers")
async def api_get_providers():
    """获取所有供应商。

    注意：必须脱敏 api_key —— 数据库里存的是 LLM 服务商的原始 key，
    直接 dict(p) 送给前端会让任何能调到这个接口的人拿到完整 key。
    前端 admin-panel 读的是 api_key_preview, 这里把 api_key 替换成
    preview 字段, 原始 key 不出后端。
    """
    try:
        providers = await get_all_providers()
        result = []
        for p in providers:
            sp = dict(p)
            raw_key = sp.pop("api_key", "") or ""
            if raw_key:
                if len(raw_key) > 12:
                    sp["api_key_preview"] = raw_key[:6] + "…" + raw_key[-4:]
                else:
                    sp["api_key_preview"] = "•" * min(len(raw_key), 8)
            else:
                sp["api_key_preview"] = ""
            if sp.get("created_at"):
                sp["created_at"] = sp["created_at"].isoformat()
            if sp.get("updated_at"):
                sp["updated_at"] = sp["updated_at"].isoformat()
            result.append(sp)
        return {"status": "ok", "providers": result}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/providers")
async def api_create_provider(request: Request):
    """创建供应商"""
    try:
        data = await request.json()
        name = data.get("name", "").strip()
        api_base_url = data.get("api_base_url", "").strip()
        api_key = data.get("api_key", "").strip()
        enabled = data.get("enabled", True)

        if not name:
            return {"error": "供应商名称不能为空"}
        if not api_base_url:
            return {"error": "API Base URL 不能为空"}

        provider = await create_provider(name, api_base_url, api_key, enabled)
        return {"status": "created", "provider": provider}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/providers/{provider_id}")
async def api_update_provider(provider_id: int, request: Request):
    """更新供应商"""
    try:
        data = await request.json()
        provider = await update_provider(provider_id, **data)
        if provider:
            return {"status": "updated", "provider": provider}
        return {"error": "供应商不存在"}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/providers/{provider_id}")
async def api_delete_provider(provider_id: int):
    """删除供应商"""
    try:
        success = await delete_provider(provider_id)
        if success:
            return {"status": "deleted"}
        return {"error": "供应商不存在"}
    except Exception as e:
        return {"error": str(e)}


def _detect_provider_type(api_base_url: str) -> str:
    """根据供应商 URL 判断类型"""
    url = (api_base_url or '').lower()
    if 'aihubmix' in url:
        return 'aihubmix'
    elif 'openrouter' in url:
        return 'openrouter'
    return 'generic'


def _transform_aihubmix_model(m: dict) -> dict:
    """将 AIHubMix 新 API 格式转换为 OpenRouter 兼容格式，前端无需改动"""
    features = [f.strip() for f in (m.get('features') or '').split(',') if f.strip()]
    input_mods = [x.strip() for x in (m.get('input_modalities') or 'text').split(',') if x.strip()]
    model_type = (m.get('types') or 'llm').strip()

    # 映射 features → supported_parameters
    params = []
    if 'thinking' in features:
        params.append('reasoning')
    if 'tools' in features or 'function_calling' in features:
        params.append('tools')
    if 'web' in features:
        params.append('web')

    # 推断 output_modalities
    output_mods = ['text']
    if model_type == 'image_generation':
        output_mods = ['image']
    elif model_type == 'video':
        output_mods = ['video']

    pricing = m.get('pricing') or {}
    # 只有真正有定价数据时才转换，避免无定价模型被误判为免费
    transformed_pricing = {}
    if pricing and (pricing.get('input') is not None or pricing.get('output') is not None):
        transformed_pricing = {
            'prompt': str(pricing.get('input', '')),
            'completion': str(pricing.get('output', '')),
        }

    return {
        'id': m.get('model_id', ''),
        'name': m.get('model_id', ''),
        'description': m.get('desc', ''),
        'architecture': {
            'input_modalities': input_mods,
            'output_modalities': output_mods,
        },
        'supported_parameters': params,
        'context_length': m.get('context_length'),
        'max_output': m.get('max_output'),
        'pricing': transformed_pricing if transformed_pricing else None,
        '_is_embedding': model_type == 'embedding',
        '_is_rerank': model_type == 'rerank',
        '_ahm_type': model_type,       # 原始类型，供前端筛选
    }


@app.get("/admin/providers/{provider_id}/models")
async def api_get_provider_models(provider_id: int):
    """从供应商 API 拉取模型列表（代理，避免前端跨域）。同时拉取聊天模型和嵌入模型。"""
    try:
        provider = await get_provider(provider_id)
        if not provider:
            return {"error": "供应商不存在"}

        # 构造基础地址
        base = provider['api_base_url'].rstrip('/')
        # 如果 base 以 /chat/completions 结尾，去掉
        if base.endswith('/chat/completions'):
            base = base.rsplit('/chat/completions', 1)[0]

        provider_type = _detect_provider_type(base)

        headers = {"Content-Type": "application/json"}
        if provider['api_key']:
            headers["Authorization"] = f"Bearer {provider['api_key']}"

        import httpx
        async with httpx.AsyncClient(timeout=30) as client:

            # ── AIHubMix：优先用新 API，失败降级旧 API ──
            if provider_type == 'aihubmix':
                # 新 API 地址：https://aihubmix.com/api/v1/models
                new_api_base = base.split('/v1')[0] if '/v1' in base else base
                try:
                    resp = await client.get(f"{new_api_base}/api/v1/models", headers=headers)
                    if resp.status_code == 200:
                        raw_models = resp.json().get("data", [])
                        models = [_transform_aihubmix_model(m) for m in raw_models]
                    else:
                        raise ValueError(f"新 API 返回 {resp.status_code}")
                except Exception:
                    # 降级到旧 /v1/models 接口
                    resp = await client.get(f"{base}/models", headers=headers)
                    if resp.status_code != 200:
                        return {"error": f"供应商返回 {resp.status_code}", "detail": resp.text[:500]}
                    models = resp.json().get("data", [])
                    provider_type = 'generic'  # 降级后按通用处理

            # ── OpenRouter / 通用：走旧 /models 接口 ──
            else:
                resp = await client.get(f"{base}/models", headers=headers)
                if resp.status_code != 200:
                    return {"error": f"供应商返回 {resp.status_code}", "detail": resp.text[:500]}
                chat_models = resp.json().get("data", [])

                # 尝试拉取嵌入模型（不是所有供应商都支持，失败不影响）
                embed_models = []
                try:
                    embed_resp = await client.get(f"{base}/embeddings/models", headers=headers)
                    if embed_resp.status_code == 200:
                        embed_data = embed_resp.json().get("data", [])
                        chat_ids = {m.get("id") for m in chat_models}
                        for m in embed_data:
                            if m.get("id") not in chat_ids:
                                m["_is_embedding"] = True
                                embed_models.append(m)
                except Exception:
                    pass

                models = chat_models + embed_models

        return {
            "status": "ok",
            "provider_id": provider_id,
            "provider_name": provider['name'],
            "provider_type": provider_type,
            "count": len(models),
            "models": models,
        }
    except httpx.TimeoutException:
        return {"error": "请求超时，请检查供应商地址"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 供应商已保存模型管理 API
# ============================================================

@app.get("/admin/all-saved-models")
async def api_get_all_saved_models():
    """获取所有供应商的已保存模型（含供应商名称，用于默认模型选择器）"""
    try:
        models = await get_all_saved_models()
        return {"status": "ok", "models": models}
    except Exception as e:
        return {"error": str(e)}


@app.get("/admin/providers/{provider_id}/saved-models")
async def api_get_saved_models(provider_id: int):
    """获取供应商已保存的模型列表"""
    try:
        models = await get_provider_models(provider_id)
        return {"status": "ok", "models": models}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/providers/{provider_id}/saved-models")
async def api_add_saved_model(provider_id: int, request: Request):
    """添加模型到供应商"""
    try:
        data = await request.json()
        model_id = data.get("model_id", "").strip()
        if not model_id:
            return {"error": "model_id 不能为空"}

        model = await add_provider_model(
            provider_id=provider_id,
            model_id=model_id,
            display_name=data.get("display_name", ""),
            model_type=data.get("model_type", "chat"),
            input_modes=data.get("input_modes", "text"),
            output_modes=data.get("output_modes", "text"),
            capabilities=data.get("capabilities", ""),
        )
        if model:
            return {"status": "created", "model": model}
        return {"error": "模型已存在"}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/saved-models/{model_pk_id}")
async def api_update_saved_model(model_pk_id: int, request: Request):
    """更新已保存模型的配置"""
    try:
        data = await request.json()
        model = await update_provider_model(model_pk_id, **data)
        if model:
            return {"status": "updated", "model": model}
        return {"error": "模型不存在"}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/saved-models/{model_pk_id}")
async def api_delete_saved_model(model_pk_id: int):
    """删除已保存的模型"""
    try:
        success = await delete_provider_model(model_pk_id)
        if success:
            return {"status": "deleted"}
        return {"error": "模型不存在"}
    except Exception as e:
        return {"error": str(e)}

        
# ============================================================
# 记忆分类管理 API（v3.7）
# ============================================================

@app.get("/admin/categories")
async def api_get_categories():
    """获取所有分类（含记忆计数）"""
    try:
        categories = await get_all_categories()
        return {"status": "ok", "categories": categories}
    except Exception as e:
        return {"error": str(e)}


@app.post("/admin/categories")
async def api_create_category(request: Request):
    """创建分类"""
    try:
        data = await request.json()
        name = data.get("name", "").strip()
        if not name:
            return {"error": "分类名称不能为空"}
        category = await create_category(
            name=name,
            color=data.get("color", "#6B7280"),
            icon=data.get("icon", "📁"),
            sort_order=data.get("sort_order", 0),
        )
        return {"status": "created", "category": category}
    except Exception as e:
        if "unique" in str(e).lower():
            return {"error": "分类名称已存在"}
        return {"error": str(e)}


@app.put("/admin/categories/{category_id}")
async def api_update_category(category_id: int, request: Request):
    """更新分类"""
    try:
        data = await request.json()
        category = await update_category(category_id, **data)
        if category:
            return {"status": "updated", "category": category}
        return {"error": "分类不存在"}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/admin/categories/{category_id}")
async def api_delete_category(category_id: int):
    """删除分类"""
    try:
        success = await delete_category(category_id)
        if success:
            return {"status": "deleted"}
        return {"error": "分类不存在"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 联网搜索 API（v3.8）
# ============================================================

@app.get("/admin/search-engines")
async def api_get_search_engines():
    """获取所有支持的搜索引擎列表"""
    return {"engines": get_engine_list()}


@app.get("/admin/search-config")
async def api_get_search_config():
    """获取当前搜索配置"""
    engine = await get_config("search_engine") or ""
    api_key = await get_config("search_api_key") or ""
    max_results = await get_config_int("search_max_results", fallback=5)
    return {
        "engine": engine,
        "api_key": api_key,
        "max_results": max_results,
    }


@app.put("/admin/search-config")
async def api_set_search_config(request: Request):
    """更新搜索配置"""
    try:
        data = await request.json()
        if "engine" in data:
            await set_config("search_engine", data["engine"])
        if "api_key" in data:
            await set_config("search_api_key", data["api_key"])
        if "max_results" in data:
            await set_config("search_max_results", str(data["max_results"]))
        return {"status": "updated"}
    except Exception as e:
        return JSONResponse(status_code=400, content={"error": str(e)})


@app.post("/admin/search-test")
async def api_search_test(request: Request):
    """测试搜索（调试用）"""
    try:
        data = await request.json()
        query = data.get("query", "")
        engine = data.get("engine") or await get_config("search_engine") or ""
        api_key = data.get("api_key") or await get_config("search_api_key") or ""
        max_results = data.get("max_results", 5)
        
        if not query:
            return JSONResponse(status_code=400, content={"error": "query 不能为空"})
        if not engine:
            return JSONResponse(status_code=400, content={"error": "未配置搜索引擎"})
        
        results = await web_search(query=query, engine=engine, api_key=api_key, max_results=max_results)
        return {
            "engine": engine,
            "query": query,
            "count": len(results),
            "results": [r.to_dict() for r in results],
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ============================================================
# MCP 客户端管理 API（v3.8）
# ============================================================

@app.post("/admin/mcp/list-tools")
async def api_mcp_list_tools(request: Request):
    """获取指定 MCP 服务器的工具列表"""
    try:
        data = await request.json()
        servers = data.get("servers", [])
        if not servers:
            return {"tools": [], "tool_map": {}}
        
        openai_tools, tool_map = await get_tools_for_servers(servers)
        return {
            "count": len(openai_tools),
            "tools": [t["function"] for t in openai_tools],
            "tool_map": {k: v["server_name"] for k, v in tool_map.items()},
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/admin/mcp/clear-cache")
async def api_mcp_clear_cache(request: Request):
    """清除 MCP 工具缓存"""
    try:
        data = await request.json()
        url = data.get("url")
        clear_tool_cache(url)
        return {"status": "cleared", "url": url or "all"}
    except Exception:
        clear_tool_cache()
        return {"status": "cleared", "url": "all"}


# ============================================================
# 供应商余额查询 API（多供应商通用）
# ============================================================

async def _query_openrouter_credits(api_key: str):
    """查询 OpenRouter 余额"""
    result = {}
    async with httpx.AsyncClient(timeout=10) as client:
        resp1 = await client.get(
            "https://openrouter.ai/api/v1/auth/key",
            headers={"Authorization": f"Bearer {api_key}"},
        )
        if resp1.status_code == 200:
            d = resp1.json().get("data", {})
            result["usage"] = d.get("usage", 0)
            result["limit"] = d.get("limit")
            result["limit_remaining"] = d.get("limit_remaining")
        
        resp2 = await client.get(
            "https://openrouter.ai/api/v1/credits",
            headers={"Authorization": f"Bearer {api_key}"},
        )
        if resp2.status_code == 200:
            d2 = resp2.json().get("data", {})
            result["total_credits"] = d2.get("total_credits", 0)
            result["total_usage"] = d2.get("total_usage", 0)
            result["balance"] = round(d2.get("total_credits", 0) - d2.get("total_usage", 0), 6)
    return result


async def _query_generic_credits(base_url: str, api_key: str):
    """尝试 OpenAI 兼容的余额查询（/v1/dashboard/billing/subscription）"""
    base = base_url.rstrip("/").split("/chat/completions")[0].rstrip("/")
    # 去掉末尾的 /v1 以拿到根域名
    root = base.rsplit("/v1", 1)[0] if "/v1" in base else base
    result = {}
    async with httpx.AsyncClient(timeout=10) as client:
        # 方式1：new-api / one-api 风格的 /v1/dashboard/billing/subscription
        try:
            resp = await client.get(
                f"{root}/v1/dashboard/billing/subscription",
                headers={"Authorization": f"Bearer {api_key}"},
            )
            if resp.status_code == 200:
                d = resp.json()
                hard_limit = d.get("hard_limit_usd") or d.get("system_hard_limit_usd", 0)
                # 过滤掉 new-api 返回的"无限额度"假数字（通常是 1亿）
                if hard_limit and hard_limit < 100000:
                    result["total_credits"] = hard_limit
        except Exception:
            pass
        
        # 方式2：/v1/dashboard/billing/usage
        try:
            from datetime import datetime, timedelta
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            start = (datetime.now(timezone.utc) - timedelta(days=90)).strftime("%Y-%m-%d")
            resp2 = await client.get(
                f"{root}/v1/dashboard/billing/usage?start_date={start}&end_date={today}",
                headers={"Authorization": f"Bearer {api_key}"},
            )
            if resp2.status_code == 200:
                d2 = resp2.json()
                usage_cents = d2.get("total_usage", 0)
                result["total_usage"] = round(usage_cents / 100, 6) if usage_cents > 1 else usage_cents
        except Exception:
            pass
        
        if "total_credits" in result:
            total_usage = result.get("total_usage", 0)
            result["balance"] = round(result["total_credits"] - total_usage, 6)
    
    return result


@app.get("/admin/credits")
async def api_get_credits():
    """查询所有已启用供应商的余额"""
    try:
        providers = await get_all_providers()
        enabled = [p for p in providers if p.get("enabled")]
        
        if not enabled:
            # 没有配置供应商，用全局环境变量兜底（向后兼容）
            if API_KEY and "openrouter" in API_BASE_URL:
                result = await _query_openrouter_credits(API_KEY)
                if result:
                    result["provider_name"] = "OpenRouter"
                    return {"providers": [result]}
            return {"providers": []}
        
        results = []
        for p in enabled:
            base = p.get("api_base_url", "")
            key = p.get("api_key", "")
            if not key:
                continue
            
            entry = {"provider_id": p["id"], "provider_name": p["name"]}
            
            if "openrouter" in base.lower():
                data = await _query_openrouter_credits(key)
            else:
                data = await _query_generic_credits(base, key)
            
            entry.update(data)
            if data:  # 只返回有数据的
                results.append(entry)
        
        return {"providers": results}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# System Prompt 管理 API（v3.7）
# ============================================================

@app.get("/admin/system-prompt")
async def api_get_system_prompt():
    """获取当前 system prompt"""
    try:
        prompt = await get_active_system_prompt()
        # 判断来源
        db_prompt = await get_system_prompt_from_db()
        source = "database" if db_prompt is not None else "file"
        return {"status": "ok", "content": prompt, "source": source, "length": len(prompt)}
    except Exception as e:
        return {"error": str(e)}


@app.put("/admin/system-prompt")
async def api_set_system_prompt(request: Request):
    """保存 system prompt 到数据库"""
    try:
        data = await request.json()
        content = data.get("content", "")
        await set_system_prompt_in_db(content)
        return {"status": "updated", "length": len(content)}
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 用户画像 API
# ============================================================

@app.post("/admin/update-profile-now")
async def api_update_profile_now():
    """手动触发用户画像更新"""
    try:
        from daily_digest import update_user_profile
        result = await update_user_profile()
        return result
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# v5.8：对话搜索 API
# ============================================================

@app.get("/search/messages")
async def api_search_messages(q: str = "", project_id: str = None, limit: int = 20):
    """
    搜索对话消息内容和标题。
    
    参数：
    - q: 搜索关键词
    - project_id: 项目ID过滤（'none' 表示只搜无项目的对话）
    - limit: 最多返回多少条匹配
    """
    try:
        from database import search_chat_messages
        results = await search_chat_messages(q, project_id=project_id, limit=limit)
        return results
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ============================================================
# 云端同步 API（v4.1）
# ============================================================

# ──── 对话 ────

@app.get("/sync/conversations")
async def api_sync_get_conversations():
    """获取对话列表（不含消息体）"""
    try:
        convs = await sync_get_conversations()
        return {"conversations": [_serialize_datetimes(c) for c in convs]}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/sync/conversations/{conv_id}")
async def api_sync_get_conversation(conv_id: str):
    """获取单个对话 + 全部消息"""
    try:
        conv = await sync_get_conversation(conv_id)
        if not conv:
            return JSONResponse(status_code=404, content={"error": "对话不存在"})
        # datetime 对象需要序列化为 ISO 字符串，否则 JSONResponse 会崩溃
        return _serialize_datetimes(conv)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/sync/conversations/{conv_id}")
async def api_sync_upsert_conversation(conv_id: str, request: Request):
    """创建或更新对话（含消息）"""
    try:
        data = await request.json()
        data["id"] = conv_id
        messages = data.pop("messages", None)
        await sync_upsert_conversation(data)
        if messages is not None:
            await sync_upsert_messages(conv_id, messages)
        return {"status": "ok"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/sync/conversations/{conv_id}")
async def api_sync_delete_conversation(conv_id: str):
    """删除对话"""
    try:
        deleted = await sync_delete_conversation(conv_id)
        return {"deleted": deleted}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 项目 ────

@app.get("/sync/projects")
async def api_sync_get_projects():
    """获取所有项目"""
    try:
        projs = await sync_get_projects()
        return {"projects": [_serialize_datetimes(p) for p in projs]}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/sync/projects/{proj_id}")
async def api_sync_upsert_project(proj_id: str, request: Request):
    """创建或更新项目"""
    try:
        data = await request.json()
        data["id"] = proj_id
        await sync_upsert_project(data)
        return {"status": "ok"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/sync/projects/{proj_id}")
async def api_sync_delete_project(proj_id: str):
    """删除项目"""
    try:
        deleted = await sync_delete_project(proj_id)
        # v5.8：删除项目时清理文件块
        try:
            from database import delete_all_file_chunks
            await delete_all_file_chunks(proj_id)
        except Exception:
            pass
        return {"deleted": deleted}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── v5.8：项目文件分块处理 ────

@app.post("/projects/{proj_id}/files/{file_id}/process")
async def api_process_file_chunks(proj_id: str, file_id: str, request: Request):
    """
    接收文件文本内容，分块 + 生成嵌入 + 存入数据库。
    前端上传文件后调用。
    body: { "file_name": "xxx.txt", "text_content": "..." }
    """
    try:
        data = await request.json()
        file_name = data.get("file_name", "")
        text_content = data.get("text_content", "")
        
        if not text_content or not text_content.strip():
            return {"chunks": 0, "message": "无文本内容"}
        
        from database import save_file_chunks, delete_file_chunks
        # 先删除旧的块（如果文件重新上传）
        await delete_file_chunks(proj_id, file_id)
        # 分块 + 嵌入 + 存储
        count = await save_file_chunks(proj_id, file_id, file_name, text_content)
        return {"chunks": count, "file_id": file_id, "file_name": file_name}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/projects/{proj_id}/files/{file_id}/chunks")
async def api_delete_file_chunks(proj_id: str, file_id: str):
    """删除某个文件的所有块"""
    try:
        from database import delete_file_chunks
        count = await delete_file_chunks(proj_id, file_id)
        return {"deleted": count}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 批量导入（localStorage → 数据库） ────

@app.post("/sync/import")
async def api_sync_import(request: Request):
    """一次性导入所有 localStorage 数据"""
    try:
        data = await request.json()
        conversations = data.get("conversations", [])
        projects = data.get("projects", [])
        result = await sync_import_all(conversations, projects)
        print(f"📦 云端同步导入完成：{result}")
        return {"status": "ok", **result}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 用户/助手配置同步（复用 config 表） ────

@app.get("/sync/settings")
async def api_sync_get_settings():
    """获取所有同步配置（头像、昵称、助手设置等）"""
    sync_keys = [
        "user_avatar", "user_nickname", "assistant_avatar", "assistant_settings",
        "custom_skills", "quick_phrases", "mcp_switches", "theme_preference",
    ]
    result = {}
    for key in sync_keys:
        val = await get_config(key)
        result[key] = val or ""
    return result


@app.put("/sync/settings")
async def api_sync_put_settings(request: Request):
    """批量更新同步配置"""
    try:
        data = await request.json()
        updated = []
        for key, value in data.items():
            ok = await set_config(key, str(value) if value is not None else "")
            if ok:
                updated.append(key)
        return {"status": "ok", "updated": updated}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 数据导出（备份 zip） ────

@app.get("/sync/export")
async def api_sync_export():
    """导出全部数据为 zip"""
    import io
    import zipfile
    from datetime import datetime, timezone

    try:
        # 收集所有数据
        convs_raw = await sync_get_conversations()
        # 为每个对话加载消息
        convs_full = []
        for c in convs_raw:
            full = await sync_get_conversation(c["id"])
            if full:
                # datetime 转 ISO 字符串
                convs_full.append(_serialize_datetimes(full))
            else:
                convs_full.append(_serialize_datetimes(c))

        projs_raw = await sync_get_projects()
        projs = [_serialize_datetimes(p) for p in projs_raw]

        # 记忆
        pool = await get_pool()
        async with pool.acquire() as conn:
            mem_rows = await conn.fetch("SELECT id, content, importance, title, memory_type, source, category_id, created_at FROM memories ORDER BY created_at DESC")
        memories = [_serialize_datetimes(dict(r)) for r in mem_rows]

        # 配置
        all_config = await get_all_config()
        config_flat = {}
        for k, v in all_config.items():
            config_flat[k] = v.get("value", "") if isinstance(v, dict) else v

        # 同步设置
        sync_keys = ["user_avatar", "user_nickname", "assistant_avatar", "assistant_settings",
                      "custom_skills", "quick_phrases", "mcp_switches", "theme_preference"]
        settings = {}
        for key in sync_keys:
            val = await get_config(key)
            settings[key] = val or ""

        # 打包 zip
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("conversations.json", json.dumps(convs_full, ensure_ascii=False, indent=2))
            zf.writestr("projects.json", json.dumps(projs, ensure_ascii=False, indent=2))
            zf.writestr("memories.json", json.dumps(memories, ensure_ascii=False, indent=2))
            zf.writestr("config.json", json.dumps(config_flat, ensure_ascii=False, indent=2))
            zf.writestr("settings.json", json.dumps(settings, ensure_ascii=False, indent=2))
        buf.seek(0)

        ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        filename = f"kiwi-mem-backup-{ts}.zip"

        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


def _serialize_datetimes(obj):
    """递归将 datetime 对象转为 ISO 字符串"""
    from datetime import datetime as _dt
    if isinstance(obj, dict):
        return {k: _serialize_datetimes(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_serialize_datetimes(v) for v in obj]
    elif isinstance(obj, _dt):
        return obj.isoformat()
    return obj


# ──── 数据导入（从备份 zip 恢复） ────

@app.post("/sync/import-backup")
async def api_sync_import_backup(file: UploadFile = File(...)):
    """从备份 zip 恢复数据"""
    import io
    import zipfile

    try:
        content = await file.read()
        buf = io.BytesIO(content)

        if not zipfile.is_zipfile(buf):
            return JSONResponse(status_code=400, content={"error": "不是有效的 zip 文件"})

        buf.seek(0)
        result = {"conversations": 0, "messages": 0, "projects": 0, "memories": 0, "settings": 0, "config": 0}

        with zipfile.ZipFile(buf, 'r') as zf:
            names = zf.namelist()

            # 导入项目
            if "projects.json" in names:
                projs = json.loads(zf.read("projects.json"))
                for p in projs:
                    await sync_upsert_project(p)
                    result["projects"] += 1

            # 导入对话 + 消息
            if "conversations.json" in names:
                convs = json.loads(zf.read("conversations.json"))
                for conv in convs:
                    messages = conv.pop("messages", [])
                    await sync_upsert_conversation(conv)
                    if messages:
                        await sync_upsert_messages(conv["id"], messages)
                        result["messages"] += len(messages)
                    result["conversations"] += 1

            # 导入记忆
            if "memories.json" in names:
                mems = json.loads(zf.read("memories.json"))
                for mem in mems:
                    try:
                        await save_memory(
                            content=mem.get("content", ""),
                            importance=mem.get("importance", 5),
                            title=mem.get("title", ""),
                            category_id=mem.get("category_id"),
                            source=mem.get("source", "backup_import"),
                        )
                        result["memories"] += 1
                    except Exception:
                        pass  # 跳过重复或无效记忆

            # 导入同步设置
            if "settings.json" in names:
                settings = json.loads(zf.read("settings.json"))
                for key, val in settings.items():
                    if val:
                        await set_config(key, str(val))
                        result["settings"] += 1

            # 导入 gateway 配置
            if "config.json" in names:
                config = json.loads(zf.read("config.json"))
                for key, val in config.items():
                    if val:
                        ok = await set_config(key, str(val))
                        if ok:
                            result["config"] += 1

        print(f"📦 备份导入完成：{result}")
        return {"status": "ok", **result}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


# ──── 数据重置 ────

@app.delete("/sync/reset")
async def api_sync_reset(request: Request):
    """重置全部聊天数据（对话+项目+同步设置），记忆和 gateway 配置保留"""
    try:
        data = await request.json()
        confirm = data.get("confirm")
        if confirm != "RESET_ALL_DATA":
            return JSONResponse(status_code=400, content={"error": "需要确认码 confirm='RESET_ALL_DATA'"})

        pool = await get_pool()
        async with pool.acquire() as conn:
            # 删除所有消息和对话（级联）
            deleted_convs = await conn.execute("DELETE FROM chat_conversations")
            deleted_projs = await conn.execute("DELETE FROM chat_projects")

            # 清除同步设置
            sync_keys = ["user_avatar", "user_nickname", "assistant_avatar", "assistant_settings",
                          "custom_skills", "quick_phrases", "mcp_switches", "theme_preference"]
            for key in sync_keys:
                await conn.execute("DELETE FROM gateway_config WHERE key = $1", key)

        print("⚠️ 数据重置完成")
        return {
            "status": "ok",
            "message": "所有聊天数据已重置",
            "deleted_conversations": deleted_convs,
            "deleted_projects": deleted_projs,
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ============================================================
# 提醒系统 API（v4.2）
# ============================================================

@app.get("/reminders")
async def api_get_reminders(all: bool = False):
    """获取提醒列表（默认只返回活跃的）"""
    try:
        reminders = await get_reminders(include_completed=all)
        return JSONResponse(content=reminders)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/reminders")
async def api_create_reminder(request: Request):
    """手动创建提醒"""
    try:
        body = await request.json()
        result = await create_reminder(body)
        return JSONResponse(content=result)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

# 注意：/reminders/due 和 /reminders/{rid}/fire 必须在 /reminders/{rid} 之前定义，
# 否则 "due" 和 "xxx/fire" 会被 {rid} 路径参数捕获

@app.get("/reminders/due")
async def api_get_due_reminders():
    """获取所有到期的提醒（前端轮询用）"""
    try:
        due = await get_due_reminders()
        return JSONResponse(content=due)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/reminders/{rid}/fire")
async def api_fire_reminder(rid: str):
    """标记提醒已触发（前端调用）"""
    try:
        reminders = await get_reminders(include_completed=True)
        reminder = next((r for r in reminders if r["id"] == rid), None)
        if not reminder:
            return JSONResponse(status_code=404, content={"error": "提醒不存在"})
        ok = await fire_reminder(rid, reminder.get("repeat_type", "once"), reminder.get("repeat_config"))
        return JSONResponse(content={"ok": ok, "repeat_type": reminder.get("repeat_type")})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.put("/reminders/{rid}")
async def api_update_reminder(rid: str, request: Request):
    """更新提醒"""
    try:
        body = await request.json()
        ok = await update_reminder(rid, body)
        if ok:
            return JSONResponse(content={"ok": True})
        return JSONResponse(status_code=404, content={"error": "提醒不存在"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.delete("/reminders/{rid}")
async def api_delete_reminder(rid: str):
    """删除提醒"""
    try:
        ok = await delete_reminder(rid)
        if ok:
            return JSONResponse(content={"ok": True})
        return JSONResponse(status_code=404, content={"error": "提醒不存在"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ============================================================
# 挂载 MCP Server（Streamable HTTP）
# ============================================================
#
# 记忆系统：/memory/mcp
#   工具：search_memory, save_memory, get_recent, trigger_digest

app.mount("/memory", get_mcp_app())
app.mount("/calendar", get_calendar_mcp_app())


# ============================================================
# 启动入口
# ============================================================

if __name__ == "__main__":
    import uvicorn
    print(f"🚀 AI Memory Gateway 启动中... 端口 {PORT}")
    print(f"📝 人设长度：{len(SYSTEM_PROMPT)} 字符")
    print(f"🤖 默认模型：{DEFAULT_MODEL}")
    print(f"🔗 API 地址：{API_BASE_URL}")
    print(f"🧠 记忆系统：{'开启' if MEMORY_ENABLED else '关闭'}")
    if MEMORY_ENABLED:
        print(f"📊 记忆提取间隔：每 {MEMORY_EXTRACT_INTERVAL} 轮")
    uvicorn.run(app, host="0.0.0.0", port=PORT)
